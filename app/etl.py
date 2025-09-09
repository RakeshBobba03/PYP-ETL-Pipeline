import os
import re
import csv
import openpyxl
import requests
import html
import zipfile
import math
from sqlalchemy.exc import SQLAlchemyError
from rapidfuzz import process, fuzz, utils
from flask import current_app
from app import db
from app.models import MemberSubmission, Member, NewItem, MatchReview

def open_csv_with_encoding_detection(file_path, mode='r'):
    """
    Open a CSV file with automatic encoding detection.
    Tries multiple encodings to handle various CSV file formats.
    Returns the file handle and the encoding used.
    """
    encodings_to_try = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1']
    
    for encoding in encodings_to_try:
        try:
            f = open(file_path, mode=mode, encoding=encoding, newline='')
            # Test read a small amount to verify encoding works
            f.seek(0)
            f.read(1024)
            f.seek(0)
            return f, encoding
        except UnicodeDecodeError:
            if 'f' in locals():
                f.close()
            continue
    
    raise Exception(f"Could not read CSV file with any of the attempted encodings: {encodings_to_try}")

BATCH_SIZE = 1000
# Configurable thresholds - Made more strict to prevent incorrect matches
FUZZY_MATCH_THRESHOLD = float(os.getenv('FUZZY_MATCH_THRESHOLD', '85.0'))  # Increased from 80% to 85%
AUTO_RESOLVE_THRESHOLD = float(os.getenv('AUTO_RESOLVE_THRESHOLD', '97.0'))  # Increased from 95% to 97%
AUTO_REJECT_THRESHOLD = float(os.getenv('AUTO_REJECT_THRESHOLD', '50.0'))  # Auto-reject if score below this threshold

# Enhanced scoring penalty configuration
LENGTH_PENALTY_MULTIPLIER = float(os.getenv('LENGTH_PENALTY_MULTIPLIER', '30.0'))  # Length difference penalty
WORD_COUNT_PENALTY_MULTIPLIER = float(os.getenv('WORD_COUNT_PENALTY_MULTIPLIER', '10.0'))  # Word count difference penalty
DIETARY_TERMS_PENALTY = float(os.getenv('DIETARY_TERMS_PENALTY', '20.0'))  # Dietary terms mismatch penalty
SPECIAL_CHARS_PENALTY = float(os.getenv('SPECIAL_CHARS_PENALTY', '15.0'))  # Special characters mismatch penalty
NUMBERS_PENALTY = float(os.getenv('NUMBERS_PENALTY', '15.0'))  # Numbers mismatch penalty
ALGORITHM_DISAGREEMENT_PENALTY = float(os.getenv('ALGORITHM_DISAGREEMENT_PENALTY', '15.0'))  # Algorithm disagreement penalty
ALGORITHM_DISAGREEMENT_THRESHOLD = float(os.getenv('ALGORITHM_DISAGREEMENT_THRESHOLD', '20.0'))  # Threshold for algorithm disagreement

# Schema field mappings for header validation
MEMBER_SCHEMA_FIELDS = {
    # Core member identification
    'memberID': ['memberid', 'member_id', 'uid', 'id', 'member uid'],
    'businessName': ['businessname', 'business_name', 'company_name', 'company', 'name', 'business', 'organization'],
    'contactFullName': ['contactfullname', 'contact_full_name', 'full_name', 'contact_name', 'name'],
    'firstName': ['firstname', 'first_name', 'fname', 'given_name'],
    'lastName': ['lastname', 'last_name', 'lname', 'family_name', 'surname'],
    'contactRole': ['contactrole', 'contact_role', 'role', 'position', 'title', 'job_title'],
    'contactEmail': ['contactemail', 'contact_email', 'email', 'e-mail', 'mail', 'contact', 'primary_email'],
    'phone': ['phone', 'telephone', 'phone_number', 'contact_phone', 'mobile', 'cell'],
    'profileImage': ['profileimage', 'profile_image', 'image', 'photo', 'avatar', 'logo'],
    
    # Address fields
    'streetAddress1': ['streetaddress1', 'street_address1', 'address', 'street', 'address1', 'street1', 'street_address'],
    'city1': ['city1', 'city', 'city_1', 'town', 'municipality'],
    'stateOrProvince1': ['stateorprovince1', 'state_or_province1', 'state', 'province', 'region', 'state1'],
    'zipCode1': ['zipcode1', 'zip_code1', 'zip', 'postal_code', 'zipcode', 'postal'],
    'country1': ['country1', 'country', 'country_1', 'nation', 'location'],
    'multipleLocations': ['multiplelocations', 'multiple_locations', 'has_multiple_locations', 'multi_location'],
    
    # System fields
    'createdAt': ['createdat', 'created_at', 'date_created', 'created_date', 'timestamp'],
    'networkStatus': ['networkstatus', 'network_status', 'status', 'network'],
    'dataSource': ['datasource', 'data_source', 'source', 'import_source'],
    'sourceFile': ['sourcefile', 'source_file', 'file_name', 'filename', 'import_file'],
    'membershipStatus': ['membershipstatus', 'membership_status', 'member_status', 'status'],
    'subscriptionStatus': ['subscriptionstatus', 'subscription_status', 'subscription'],
    'isTrial': ['istrial', 'is_trial', 'trial', 'trial_member', 'trial_status'],
    
    # Offerings and services
    'memberOfferings': ['memberofferings', 'member_offerings', 'offerings', 'services_offered'],
    'designServices': ['designservices', 'design_services', 'design'],
    'suppliedEquipment': ['suppliedequipment', 'supplied_equipment', 'equipment_supplied'],
    'facilityEquipment': ['facilityequipment', 'facility_equipment', 'equipment_facility'],
    'ingredients': ['ingredients', 'ingredient', 'ingredient_list', 'components', 'materials'],
    'laboratoryServices': ['laboratoryservices', 'laboratory_services', 'lab_services', 'labservices'],
    'legalServices': ['legalservices', 'legal_services', 'legal'],
    'logisticalServices': ['logisticalservices', 'logistical_services', 'logistics', 'logistics_services'],
    'marketingServices': ['marketingservices', 'marketing_services', 'marketing'],
    'deliveredIn': ['deliveredin', 'delivered_in', 'delivery_packaging', 'packaging_delivered'],
    'suppliedPackaging': ['suppliedpackaging', 'supplied_packaging', 'packaging_supplied'],
    'regulatoryServices': ['regulatoryservices', 'regulatory_services', 'regulatory'],
    'manufacturingServices': ['manufacturingservices', 'manufacturing_services', 'manufacturing'],
    'startupFriendlyServices': ['startupfriendlyservices', 'startup_friendly_services', 'startup_services', 'startupservices'],
    'facilityDetails': ['facilitydetails', 'facility_details', 'facility'],
    'facilityAmenities': ['facilityamenities', 'facility_amenities', 'amenities'],
    'typeOfSpace': ['typeofspace', 'type_of_space', 'space_type', 'spaces'],
    'typeOfAgreement': ['typeofagreement', 'type_of_agreement', 'agreement_type', 'agreements'],
    'consultingServices': ['consultingservices', 'consulting_services', 'consulting'],
    
    # Social media and web presence
    'website': ['website', 'web', 'site', 'url', 'web_site'],
    'facebookURL': ['facebookurl', 'facebook_url', 'facebook', 'fb_url'],
    'instagramURL': ['instagramurl', 'instagram_url', 'instagram', 'ig_url'],
    'linkedinURL': ['linkedinurl', 'linkedin_url', 'linkedin', 'li_url'],
    'twitterURL': ['twitterurl', 'twitter_url', 'twitter', 'tw_url'],
    'youtubeURL': ['youtubeurl', 'youtube_url', 'youtube', 'yt_url'],
    
    # Business information
    'companyBio': ['companybio', 'company_bio', 'bio', 'business_bio', 'description', 'about', 'company description', 'business description', 'company bio'],
    'businessDetails': ['businessdetails', 'business_details', 'business_info', 'company_details'],
    'idealClient': ['idealclient', 'ideal_client', 'target_client', 'client_profile'],
    'moq': ['moq', 'minimum_order_quantity', 'min_order', 'minimum_order'],
    'readyToStartNow': ['readytostartnow', 'ready_to_start_now', 'available_now', 'ready_now'],
    'bookingProjectsForMonth': ['bookingprojectsformonth', 'booking_projects_for_month', 'projects_month'],
    'bookingProjectsForYear': ['bookingprojectsforyear', 'booking_projects_for_year', 'projects_year'],
    'leadTimes': ['leadtimes', 'lead_times', 'lead_time', 'delivery_time'],
    
    # Products and materials
    'products': ['products', 'product', 'product_list', 'items', 'goods'],
    'certifications': ['certifications', 'certification', 'certs', 'cert_list'],
    'allergens': ['allergens', 'allergen', 'allergen_list', 'allergies'],
    'byProducts': ['byproducts', 'by_products', 'byproduct', 'by_product'],
    'upCycledIngredients': ['upcycledingredients', 'up_cycled_ingredients', 'upcycled', 'recycled_ingredients'],
    
    # Sustainability and special fields
    'sustainability': ['sustainability', 'sustainable', 'sustainability_info', 'eco_friendly'],
    
    # Additional fields that might be present
    'firebaseUID': ['firebaseuid', 'firebase_uid', 'firebase_id', 'user_id'],
    'updatedAt': ['updatedat', 'updated_at', 'last_updated', 'modified_date'],
}

def get_schema_field_mapping():
    """Get the schema field mapping for header validation"""
    return MEMBER_SCHEMA_FIELDS

def is_empty_or_invalid(value):
    """
    Check if a value is empty or invalid according to requirements.
    Handles: '', 'null', 'none', 'n/a', NaN (actual NaN values)
    """
    if value is None:
        return True
    
    # Handle actual NaN values (from pandas/Excel)
    if isinstance(value, float) and math.isnan(value):
        return True
    
    # Handle string values
    if isinstance(value, str):
        normalized = value.strip().lower()
        return normalized in ('', 'null', 'none', 'n/a', 'na', 'nan')
    
    return False

def fetch_member_offerings_from_dgraph():
    """
    Fetch member offerings from Dgraph to get the actual UIDs for the playground environment.
    Returns a mapping of offering titles to their UIDs.
    """
    try:
        url = current_app.config.get('DGRAPH_URL')
        token = current_app.config.get('DGRAPH_API_TOKEN')
    except RuntimeError:
        # Not in Flask context
        return {}
    
    if not url or not token:
        try:
            current_app.logger.warning("[member_offerings] Dgraph not configured - using fallback UIDs")
        except RuntimeError:
            pass  # Not in Flask context
        return {}
    
    gql = """
    query {
      memberOfferings: queryMemberOffering {
        title
        offeringID
      }
    }
    """
    
    try:
        try:
            current_app.logger.info("[member_offerings] Fetching member offerings from Dgraph...")
        except RuntimeError:
            pass  # Not in Flask context
            
        resp = requests.post(
            url, 
            json={"query": gql}, 
            headers={"Content-Type": "application/json", "Dg-Auth": token}, 
            timeout=10
        )
        resp.raise_for_status()
        resp_json = resp.json() if resp else {}
        if resp_json is None:
            resp_json = {}
        data = resp_json.get("data", {})
        offerings = data.get('memberOfferings', [])
        
        # Create mapping of title to UID
        offerings_map = {offering['title']: offering['offeringID'] for offering in offerings}
        
        try:
            current_app.logger.info(f"[member_offerings] Fetched {len(offerings)} offerings from Dgraph: {list(offerings_map.keys())}")
        except RuntimeError:
            pass  # Not in Flask context
        
        return offerings_map
        
    except Exception as e:
        try:
            current_app.logger.error(f"[member_offerings] Could not fetch offerings from Dgraph: {e}")
        except RuntimeError:
            pass  # Not in Flask context
        return {}

def get_member_offerings_mapping():
    """
    Get the mapping for member offerings based on the provided table.
    Maps source fields to offering titles and UIDs.
    Uses dynamic UIDs from Dgraph if available, otherwise falls back to production UIDs.
    """
    # Try to fetch from Dgraph first
    dgraph_offerings = fetch_member_offerings_from_dgraph()
    
    # Define the mapping with offering titles
    base_mapping = {
        'designServices': {
            'title': 'Design',
            'fallback_uid': '0x19f191'
        },
        'suppliedEquipment': {
            'title': 'Equipment',
            'fallback_uid': '0x494de'
        },
        'facilityEquipment': {
            'title': 'Equipment',
            'fallback_uid': '0x494de'
        },
        'ingredients': {
            'title': 'Ingredients',
            'fallback_uid': '0x2626b4'
        },
        'laboratoryServices': {
            'title': 'Laboratory',
            'fallback_uid': '0x928dd'
        },
        'legalServices': {
            'title': 'Legal',
            'fallback_uid': '0x2192be'
        },
        'logisticalServices': {
            'title': 'Logistics',
            'fallback_uid': '0x200c34'
        },
        'marketingServices': {
            'title': 'Marketing',
            'fallback_uid': '0x30e3a'
        },
        'deliveredIn': {
            'title': 'Packaging',
            'fallback_uid': '0x928dc'
        },
        'suppliedPackaging': {
            'title': 'Packaging',
            'fallback_uid': '0x928dc'
        },
        'regulatoryServices': {
            'title': 'Regulatory',
            'fallback_uid': '0x7a21e'
        },
        'facilityDetails': {
            'title': 'Spaces',
            'fallback_uid': '0x19f18f'
        },
        'typeOfSpace': {
            'title': 'Spaces',
            'fallback_uid': '0x19f18f'
        },
        'typeOfAgreement': {
            'title': 'Spaces',
            'fallback_uid': '0x19f18f'
        },
        'facilityAmenities': {
            'title': 'Spaces',
            'fallback_uid': '0x19f18f'
        },
        'manufacturingServices': {
            'title': 'Manufacturing',
            'fallback_uid': '0x2c411f'
        },
        'startupFriendlyServices': {
            'title': 'R&D',
            'fallback_uid': '0x19f192'
        },
        'consultingServices': {
            'title': 'Consulting',
            'fallback_uid': '0x2aba6c'
        }
    }
    
    # Build the final mapping using Dgraph UIDs if available, otherwise fallback
    final_mapping = {}
    for field_name, offering_info in base_mapping.items():
        title = offering_info['title']
        uid = dgraph_offerings.get(title, offering_info['fallback_uid'])
        
        final_mapping[field_name] = {
            'title': title,
            'uid': uid,
            'source': 'dgraph' if title in dgraph_offerings else 'fallback'
        }
    
    return final_mapping

def determine_member_offerings(member_data, mapping):
    """
    Determine member offerings based on the presence of specific fields in member data.
    Uses the mapping table to assign offerings based on what services/fields the member has.
    """
    offerings = []
    offerings_mapping = get_member_offerings_mapping()
    
    current_app.logger.info(f"[determine_member_offerings] Member data keys: {list(member_data.keys())}")
    current_app.logger.info(f"[determine_member_offerings] Mapping keys: {list(mapping.keys())}")
    current_app.logger.info(f"[determine_member_offerings] Available offerings: {list(offerings_mapping.keys())}")
    
    # Check each field that could indicate an offering
    for field_name, offering_info in offerings_mapping.items():
        current_app.logger.debug(f"[determine_member_offerings] Checking offering field: {field_name}")
        # Special case for Manufacturing: check both manufacturingServices AND presence of products
        if field_name == 'manufacturingServices':
            manufacturing_detected = False
            
            # Check for manufacturingServices field
            for csv_header, header_info in mapping.items():
                if (isinstance(header_info, dict) and 
                    header_info.get('schema_field') == field_name and
                    member_data.get(csv_header)):
                    
                    field_value = member_data.get(csv_header, '').strip()
                    if field_value and field_value.lower() not in ['', 'n/a', 'none', 'null', 'undefined']:
                        offerings.append({
                            'title': offering_info['title'],
                            'uid': offering_info['uid'],
                            'source_field': field_name,
                            'source_value': field_value
                        })
                        manufacturing_detected = True
                        break
            
            # Also check for presence of products (as per requirements: "manufacturingServices or presence of products")
            if not manufacturing_detected:
                for csv_header, header_info in mapping.items():
                    if (isinstance(header_info, dict) and 
                        header_info.get('schema_field') == 'products' and
                        member_data.get(csv_header)):
                        
                        field_value = member_data.get(csv_header, '').strip()
                        if field_value and field_value.lower() not in ['', 'n/a', 'none', 'null', 'undefined']:
                            offerings.append({
                                'title': offering_info['title'],
                                'uid': offering_info['uid'],
                                'source_field': 'products',
                                'source_value': field_value
                            })
                            break
        else:
            # Regular field checking for all other offerings
            for csv_header, header_info in mapping.items():
                if (isinstance(header_info, dict) and 
                    header_info.get('schema_field') == field_name and
                    member_data.get(csv_header)):
                    
                    # Check if the field has meaningful data
                    field_value = member_data.get(csv_header, '').strip()
                    if field_value and field_value.lower() not in ['', 'n/a', 'none', 'null', 'undefined']:
                        offerings.append({
                            'title': offering_info['title'],
                            'uid': offering_info['uid'],
                            'source_field': field_name,
                            'source_value': field_value
                        })
                        break  # Found this offering, move to next
    
    # Special case: if no offerings detected, don't add anything
    # The "No Offerings" placeholder was causing Dgraph errors because the UID doesn't exist
    # Instead, we'll just return an empty list and let the member be created without offerings
    if not offerings:
        current_app.logger.debug(f"[detect_member_offerings] No offerings detected for member, skipping offerings assignment")
    
    return offerings

def get_member_offerings_from_cache(member_id):
    """Get member offerings from the database or session cache"""
    try:
        from flask import current_app
        current_app.logger.debug(f"[get_member_offerings_from_cache] Getting offerings for member ID: {member_id}")
        
        # First try to get from database
        member = Member.query.get(member_id)
        current_app.logger.debug(f"[get_member_offerings_from_cache] Member query result: {member}")
        if member:
            current_app.logger.debug(f"[get_member_offerings_from_cache] Member has member_offerings attribute: {hasattr(member, 'member_offerings')}")
            if hasattr(member, 'member_offerings'):
                current_app.logger.debug(f"[get_member_offerings_from_cache] Member.member_offerings value: {member.member_offerings}")
                if member.member_offerings:
                    current_app.logger.debug(f"[get_member_offerings_from_cache] Retrieved offerings from database for member {member_id}: {member.member_offerings}")
                    return member.member_offerings
            else:
                current_app.logger.warning(f"[get_member_offerings_from_cache] Member {member_id} does not have member_offerings attribute - database column may not exist")
        
        # Fallback to session cache
        if not hasattr(db.session, 'member_offerings_cache'):
            current_app.logger.debug(f"[get_member_offerings_from_cache] No member_offerings_cache attribute on db.session")
            return []
        
        cache = db.session.member_offerings_cache
        current_app.logger.debug(f"[get_member_offerings_from_cache] Cache object: {cache} (type: {type(cache)})")
        
        if cache is None:
            current_app.logger.debug(f"[get_member_offerings_from_cache] Cache is None")
            return []
        
        if not hasattr(cache, 'get'):
            current_app.logger.warning(f"[get_member_offerings_from_cache] Cache object has no 'get' method: {type(cache)}")
            return []
        
        result = cache.get(member_id, [])
        current_app.logger.debug(f"[get_member_offerings_from_cache] Retrieved offerings from cache for member {member_id}: {result} (type: {type(result)})")
        return result
        
    except Exception as e:
        try:
            from flask import current_app
            current_app.logger.error(f"[get_member_offerings_from_cache] ERROR getting offerings for member {member_id}: {e}", exc_info=True)
        except:
            pass  # In case we're not in Flask context
        return []

def map_headers_to_schema(headers):
    """
    Automatically map incoming headers to Member schema fields.
    Returns a mapping of incoming header -> schema field and any unmapped headers.
    """
    if not headers:
        return {}, []
    
    # Normalize headers for comparison
    normalized_headers = [h.strip().lower() if h else '' for h in headers]
    
    # Debug logging
    try:
        from flask import current_app
        current_app.logger.info(f"[map_headers_to_schema] Headers: {headers}")
        current_app.logger.info(f"[map_headers_to_schema] Normalized headers: {normalized_headers}")
    except:
        pass  # In case we're not in Flask context
    
    mapping = {}
    unmapped = []
    
    for header, normalized_header in zip(headers, normalized_headers):
        if not normalized_header:
            unmapped.append(header)
            continue
            
        # Try to find a match in schema fields
        best_match = None
        best_score = 0
        
        # Debug logging for this header
        try:
            from flask import current_app
            current_app.logger.info(f"[map_headers_to_schema] Processing header: '{header}' (normalized: '{normalized_header}')")
        except:
            pass
        
        for schema_field, variations in MEMBER_SCHEMA_FIELDS.items():
            for variation in variations:
                # Exact match
                if normalized_header == variation:
                    best_match = schema_field
                    best_score = 100
                    break
                # Fuzzy match
                score = fuzz.ratio(normalized_header, variation)
                if score > best_score and score > 60:  # Lower threshold for fuzzy matching
                    best_score = score
                    best_match = schema_field
                
                # Debug logging for fuzzy matches
                try:
                    from flask import current_app
                    if normalized_header == 'company bio':  # Only log for the problematic header
                        current_app.logger.info(f"[map_headers_to_schema] '{normalized_header}' vs '{variation}': score={score}")
                except:
                    pass
            
            if best_score == 100:  # Exact match found, no need to check other fields
                break
        
        if best_match and best_score >= 60:
            mapping[header] = {
                'schema_field': best_match,
                'confidence': best_score,
                'original_header': header
            }
            # Debug logging for successful mapping
            try:
                from flask import current_app
                current_app.logger.info(f"[map_headers_to_schema] Mapped '{header}' -> '{best_match}' (score: {best_score})")
            except:
                pass
        else:
            unmapped.append(header)
            # Debug logging for unmapped headers
            try:
                from flask import current_app
                current_app.logger.info(f"[map_headers_to_schema] Unmapped '{header}' (best score: {best_score})")
            except:
                pass
    
    return mapping, unmapped

def validate_required_columns(headers, mapping):
    """
    Validate that all required columns are present in the mapped headers.
    Returns validation result and missing columns.
    """
    # Required fields for this injection process (as specified in requirements)
    required_fields = [
        'businessName',      # Required - business name
        'contactEmail',      # Required - contact email
        'streetAddress1',    # Required - street address
        'city1',            # Required - city
        'country1',         # Required - country
        'companyBio'        # Required - company bio/business details
    ]
    
    # Optional but important fields (not required but should be flagged if missing)
    important_fields = [
        'products',         # Important - products list
        'ingredients',      # Important - ingredients list
        'website',          # Important - website
        'phone',           # Important - phone number
        'stateOrProvince1', # Important - state/province
        'zipCode1'         # Important - zip code
    ]
    
    # Get mapped schema fields
    mapped_fields = set()
    for header_info in mapping.values():
        if isinstance(header_info, dict) and 'schema_field' in header_info:
            mapped_fields.add(header_info['schema_field'])
    
    # Check for missing required fields
    missing_required = []
    for field in required_fields:
        if field not in mapped_fields:
            missing_required.append(field)
    
    # Check for missing important fields
    missing_important = []
    for field in important_fields:
        if field not in mapped_fields:
            missing_important.append(field)
    
    is_valid = len(missing_required) == 0
    
    return {
        'is_valid': is_valid,
        'missing_fields': missing_required,
        'missing_important': missing_important,
        'mapped_fields': list(mapped_fields),
        'total_headers': len(headers),
        'mapped_headers': len(mapping),
        'required_fields': required_fields,
        'important_fields': important_fields
    }

def normalize_data_sample(file_path, headers, mapping, sample_size=10):
    """
    Generate a normalized data sample for preview.
    Returns the first N rows with normalized data according to the mapping.
    """
    sample_data = []
    
    try:
        ext = file_path.lower().rsplit('.', 1)[1]
        
        if ext == 'csv':
            f, encoding = open_csv_with_encoding_detection(file_path)
            try:
                reader = csv.DictReader(f)
                for i, row in enumerate(reader):
                    if i >= sample_size:
                        break
                    
                    normalized_row = normalize_row_data(row, headers, mapping)
                    sample_data.append(normalized_row)
            finally:
                f.close()
                    
        elif ext in ['xlsx', 'xls']:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            
            # Skip header row
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if i > sample_size + 1:  # +1 because we start from row 2
                    break
                
                # Convert row to dict-like structure
                row_dict = {}
                for j, header in enumerate(headers):
                    if j < len(row):
                        row_dict[header] = row[j]
                    else:
                        row_dict[header] = None
                
                normalized_row = normalize_row_data(row_dict, headers, mapping)
                sample_data.append(normalized_row)
            
            wb.close()
    
    except Exception as e:
        current_app.logger.error(f"Error generating data sample: {e}")
        return []
    
    return sample_data

def normalize_row_data(row, headers, mapping):
    """
    Normalize a single row of data according to the header mapping.
    Returns normalized data with schema field names.
    """
    normalized = {}
    
    for header in headers:
        if header in mapping and isinstance(mapping[header], dict):
            schema_field = mapping[header]['schema_field']
            value = row.get(header, '')
            
            # Normalize the value
            if is_empty_or_invalid(value):
                value = ''
            else:
                value = str(value).strip()
            
            normalized[schema_field] = value
        else:
            # Unmapped header - store with original name
            value = row.get(header, '')
            if is_empty_or_invalid(value):
                value = ''
            else:
                value = str(value).strip()
            normalized[f"unmapped_{header}"] = value
    
    return normalized

def validate_excel_file(file_path):
    """Validate Excel file before processing to prevent zip file errors"""
    try:
        # Check if file exists and has content
        if not os.path.exists(file_path):
            return False, "File does not exist"
        
        if os.path.getsize(file_path) == 0:
            return False, "File is empty"
        
        # For .xlsx files, verify it's a valid zip archive
        if file_path.lower().endswith('.xlsx'):
            try:
                with zipfile.ZipFile(file_path, 'r') as zip_file:
                    # Check if it contains the expected Excel structure
                    file_list = zip_file.namelist()
                    if not any(name.startswith('xl/') for name in file_list):
                        return False, "File does not contain valid Excel structure"
            except zipfile.BadZipFile:
                return False, "File is not a valid Excel file (corrupted or wrong format)"
        
        # Try to open with openpyxl to verify it's readable
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if not wb.active:
            return False, "Excel file has no active sheet"
        wb.close()
        
        return True, "File is valid"
    except Exception as e:
        return False, f"File validation failed: {str(e)}"

def get_fuzzy_match_threshold():
    """Get fuzzy matching threshold from environment or use default"""
    return FUZZY_MATCH_THRESHOLD

def get_auto_resolve_threshold():
    """Get auto-resolution threshold from environment or use default"""
    return AUTO_RESOLVE_THRESHOLD

def get_auto_reject_threshold():
    """Get auto-reject threshold from environment or use default"""
    return AUTO_REJECT_THRESHOLD

def apply_match_penalties(text_sanitized, match_name, raw_score):
    """
    Apply consistent penalties to any match score based on the enhanced scoring system.
    This ensures both best matches and alternatives use the same penalty logic.
    """
    adjusted_score = raw_score
    
    # Penalty 1: Length difference penalty
    length_diff = abs(len(text_sanitized) - len(match_name))
    max_length = max(len(text_sanitized), len(match_name))
    length_penalty = (length_diff / max_length) * LENGTH_PENALTY_MULTIPLIER
    adjusted_score -= length_penalty
    
    # Penalty 2: Word count difference penalty
    text_word_count = len(text_sanitized.split())
    match_word_count = len(match_name.split())
    word_diff = abs(text_word_count - match_word_count)
    word_penalty = min(word_diff * WORD_COUNT_PENALTY_MULTIPLIER, 25)
    adjusted_score -= word_penalty
    
    # Penalty 3: Dietary terms penalty
    dietary_terms = ['gluten-free', 'organic', 'natural', 'raw', 'extra virgin', 'whole grain']
    text_has_dietary = any(term in text_sanitized.lower() for term in dietary_terms)
    match_has_dietary = any(term in match_name.lower() for term in dietary_terms)
    if text_has_dietary != match_has_dietary:
        adjusted_score -= DIETARY_TERMS_PENALTY
    
    # Penalty 4: Special characters penalty
    text_special_chars = sum(1 for c in text_sanitized if c in '!@#$%^&*()')
    match_special_chars = sum(1 for c in match_name if c in '!@#$%^&*()')
    if text_special_chars != match_special_chars:
        adjusted_score -= SPECIAL_CHARS_PENALTY
    
    # Penalty 5: Numbers penalty
    text_has_numbers = any(c.isdigit() for c in text_sanitized)
    match_has_numbers = any(c.isdigit() for c in match_name)
    if text_has_numbers != match_has_numbers:
        adjusted_score -= NUMBERS_PENALTY
    
    # Ensure score doesn't go below 0
    adjusted_score = max(0.0, adjusted_score)
    
    return adjusted_score

def sanitize_string(value):
    """Sanitize string input to prevent XSS and injection attacks"""
    if not value:
        return value
    # Remove HTML tags and escape special characters
    value = str(value).strip()
    value = re.sub(r'<[^>]+>', '', value)  # Remove HTML tags
    value = html.escape(value)  # Escape HTML entities
    return value

def normalize_offering_text(text):
    """
    Advanced normalization for offering text with capital case, punctuation handling,
    and common variant normalization.
    
    Requirements:
    - Capital Case, trim, remove punctuation (but preserve relevant characters like hyphens, periods for scientific names)
    - Normalize common variants (e.g., "vit C" → "Vitamin C")
    - Full names for new offerings (e.g., "B. adolescentis" → "Bifidobacterium adolescentis")
    """
    if not text or not isinstance(text, str):
        return text
    
    # Step 1: Basic cleanup
    text = text.strip()
    if not text:
        return text
    
    # Step 2: Collapse multiple whitespace
    text = re.sub(r'\s+', ' ', text)
    
    # Step 3: Common variant normalization (before punctuation removal)
    variant_mappings = {
        # Vitamin variants
        r'\bvit\s+c\b': 'Vitamin C',
        r'\bvit\s+d\b': 'Vitamin D',
        r'\bvit\s+b\b': 'Vitamin B',
        r'\bvit\s+e\b': 'Vitamin E',
        r'\bvit\s+a\b': 'Vitamin A',
        r'\bvit\s+k\b': 'Vitamin K',
        
        # Common abbreviations
        r'\bprobiotics?\b': 'Probiotics',
        r'\bprebiotics?\b': 'Prebiotics',
        r'\bomega\s*3\b': 'Omega-3',
        r'\bomega\s*6\b': 'Omega-6',
        r'\bomega\s*9\b': 'Omega-9',
        r'\bcoq10\b': 'CoQ10',
        r'\bco\s*q\s*10\b': 'CoQ10',
        
        # Scientific name expansions
        r'\bB\.\s*adolescentis\b': 'Bifidobacterium adolescentis',
        r'\bB\.\s*lactis\b': 'Bifidobacterium lactis',
        r'\bB\.\s*bifidum\b': 'Bifidobacterium bifidum',
        r'\bL\.\s*acidophilus\b': 'Lactobacillus acidophilus',
        r'\bL\.\s*rhamnosus\b': 'Lactobacillus rhamnosus',
        r'\bL\.\s*casei\b': 'Lactobacillus casei',
        r'\bS\.\s*boulardii\b': 'Saccharomyces boulardii',
        r'\bS\.\s*cerevisiae\b': 'Saccharomyces cerevisiae',
        
        # Common ingredient variants
        r'\bstevia\b': 'Stevia',
        r'\bmonk\s*fruit\b': 'Monk Fruit',
        r'\bmonkfruit\b': 'Monk Fruit',
        r'\bchicory\s*root\b': 'Chicory Root',
        r'\binulin\b': 'Inulin',
        r'\bpectin\b': 'Pectin',
        r'\bguar\s*gum\b': 'Guar Gum',
        r'\bxanthan\s*gum\b': 'Xanthan Gum',
        r'\bcarrageenan\b': 'Carrageenan',
        r'\bagar\b': 'Agar',
        r'\bgelatin\b': 'Gelatin',
        r'\bgelatine\b': 'Gelatin',
    }
    
    # Apply variant mappings (case-insensitive)
    for pattern, replacement in variant_mappings.items():
        text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)
    
    # Step 4: Remove punctuation but preserve scientific/chemical notation
    # Keep: hyphens, periods (for scientific names), parentheses, numbers
    # Remove: commas, semicolons, exclamation marks, question marks, quotes, etc.
    text = re.sub(r'[,\;\!\?\'\""\[\]{}]', '', text)
    
    # Step 5: Capital Case normalization
    # Split into words and capitalize each word appropriately
    words = text.split()
    normalized_words = []
    
    for word in words:
        if not word:
            continue
            
        # Handle special cases
        if word.lower() in ['and', 'or', 'of', 'the', 'in', 'on', 'at', 'to', 'for', 'with', 'by']:
            # Keep common words lowercase unless they're the first word
            if not normalized_words:
                normalized_words.append(word.capitalize())
            else:
                normalized_words.append(word.lower())
        elif '-' in word:
            # Handle hyphenated words (e.g., "Omega-3", "CoQ10")
            parts = word.split('-')
            capitalized_parts = []
            for part in parts:
                if part:
                    capitalized_parts.append(part.capitalize())
            normalized_words.append('-'.join(capitalized_parts))
        elif '.' in word and len(word) > 1:
            # Handle scientific abbreviations (e.g., "B.adolescentis")
            parts = word.split('.')
            capitalized_parts = []
            for part in parts:
                if part:
                    capitalized_parts.append(part.capitalize())
            normalized_words.append('.'.join(capitalized_parts))
        else:
            # Regular word capitalization
            normalized_words.append(word.capitalize())
    
    # Join words back together
    result = ' '.join(normalized_words)
    
    # Step 6: Final cleanup - remove extra spaces
    result = re.sub(r'\s+', ' ', result).strip()
    
    return result

def validate_business_name(name):
    """Validate business name for security and data quality"""
    if not name or len(name.strip()) < 2:
        return False, "Business name must be at least 2 characters long"
    if len(name.strip()) > 200:
        return False, "Business name must be less than 200 characters"
    if re.search(r'[<>"\']', name):  # Check for dangerous characters
        return False, "Business name contains invalid characters"
    return True, None

def validate_email(email):
    """Basic email validation"""
    if not email:
        return True, None  # Email is optional
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(email_pattern, email):
        return False, "Invalid email format"
    return True, None

def process_submission_file(filename, custom_mapping=None):
    """
    Load a new submission, validate each row, skip bad ones,
    return (count, validation_errors, valid_row_indices)
    """
    # Use database transaction to prevent race conditions
    with db.session.begin():
        # Check if already processed within transaction
        existing = MemberSubmission.query.filter_by(name=filename).with_for_update().first()
        if existing:
            current_app.logger.info(f"[etl] skipping {filename}: already processed")
            return 0, [], []

        # Create submission record immediately to prevent race conditions
        sub = MemberSubmission(name=filename)
        db.session.add(sub)
        db.session.flush()  # Get the ID but don't commit yet
        
        try:
            # Process the file
            result = _process_file_content(filename, sub, custom_mapping)
            # If we get here, processing succeeded, so commit
            return result
        except Exception as e:
            # If any error occurs, the transaction will rollback automatically
            current_app.logger.error(f"[etl] Error processing {filename}: {e}")
            raise

def _process_file_content(filename, submission, custom_mapping=None):
    """Process the actual file content (separated for better error handling)"""
    data_dir = os.path.join(os.getcwd(), 'seed_data', 'new_submissions')
    fp = os.path.join(data_dir, filename)
    ext = filename.lower().rsplit('.', 1)[1]

    # Use custom mapping if provided
    
    # Prepare reader + helper - process row by row to avoid memory issues
    if ext == 'csv':
        f, encoding = open_csv_with_encoding_detection(fp)
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        
        # Apply custom mapping if available
        if custom_mapping:
            # Create reverse mapping: schema_field -> original_header
            reverse_mapping = {v: k for k, v in custom_mapping.items()}
            current_app.logger.info(f"[etl] Custom mapping applied: {custom_mapping}")
            current_app.logger.info(f"[etl] Reverse mapping: {reverse_mapping}")
            get = lambda r, c: r.get(reverse_mapping.get(c, c))
        else:
            get = lambda r, c: r.get(c)
            
        # Process CSV row by row
        return _process_csv_rows(reader, headers, get, submission, custom_mapping)
    elif ext in ['xlsx', 'xls']:
        return _process_excel_file_safe(fp, filename, submission, custom_mapping)
    else:
        raise ValueError(f"Unsupported file format: {ext}. Only .csv, .xlsx, and .xls files are supported.")

def _process_excel_file_safe(file_path, filename, submission, custom_mapping=None):
    """Safely process Excel files with multiple fallback methods"""
    # First, try to validate the file
    is_valid, validation_msg = validate_excel_file(file_path)
    if not is_valid:
        raise ValueError(f"Excel file validation failed: {validation_msg}")
    
    try:
        # Method 1: Try with openpyxl (most reliable for .xlsx)
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active
        if not sheet:
            raise ValueError("Excel file has no active sheet")
            
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [h if h else '' for h in header_row]
        
        # Apply custom mapping if available
        if custom_mapping:
            # Create reverse mapping: schema_field -> original_header
            reverse_mapping = {v: k for k, v in custom_mapping.items()}
            current_app.logger.info(f"[etl] Excel custom mapping applied: {custom_mapping}")
            current_app.logger.info(f"[etl] Excel reverse mapping: {reverse_mapping}")
            get = lambda r, c: r[headers.index(reverse_mapping.get(c, c))] if reverse_mapping.get(c, c) in headers else None
        else:
            get = lambda r, c: r[headers.index(c)] if c in headers else None
        
        # Process Excel row by row
        result = _process_excel_rows(sheet, headers, get, submission, custom_mapping)
        wb.close()
        return result
        
    except Exception as e:
        error_msg = str(e)
        
        # Provide specific guidance based on error type
        if "File is not a zip file" in error_msg or "not a zip file" in error_msg:
            raise ValueError(
                f"The file '{filename}' cannot be read as an Excel file. This usually means:\n"
                f"1. The file is corrupted\n"
                f"2. The file was saved in an unsupported format\n"
                f"3. The file extension doesn't match its actual format\n\n"
                f"Please try:\n"
                f"• Opening the file in Excel and re-saving it as .xlsx\n"
                f"• Converting it to CSV format\n"
                f"• Checking if the file is actually an Excel file"
            )
        elif "BadZipFile" in error_msg:
            raise ValueError(
                f"The file '{filename}' appears to be corrupted or not a valid Excel file. "
                f"Please try re-saving it from Excel or convert it to CSV format."
            )
        else:
            raise ValueError(f"Error reading Excel file '{filename}': {error_msg}")

def _process_csv_rows(reader, headers, get, submission, custom_mapping=None):
    """Process CSV rows one by one to avoid memory issues"""
    return _process_rows_generator(reader, headers, get, submission, is_csv=True, custom_mapping=custom_mapping)

def _process_excel_rows(sheet, headers, get, submission, custom_mapping=None):
    """Process Excel rows one by one to avoid memory issues"""
    return _process_rows_generator(sheet.iter_rows(min_row=2, values_only=True), headers, get, submission, is_csv=False, custom_mapping=custom_mapping)

def _process_rows_generator(rows_generator, headers, get, submission, is_csv=True, custom_mapping=None):
    """Generic row processor that handles both CSV and Excel data"""
    # Must have these columns in the file
    required_columns = ['businessName', 'country1', 'contactEmail', 'streetAddress1', 'city1', 'products', 'ingredients']
    
    # Check if we have custom mapping
    if custom_mapping:
        # Use custom mapping to find required columns
        mapped_headers = set()
        for header, schema_field in custom_mapping.items():
            mapped_headers.add(schema_field)
        
        for col in required_columns:
            if col not in mapped_headers:
                current_app.logger.error(f"[etl] File missing required column: {col}")
                raise ValueError(f"Missing required column: {col}")
    else:
        # Use original header validation
        for col in required_columns:
            if col not in headers:
                current_app.logger.error(f"[etl] File missing required column: {col}")
                raise ValueError(f"Missing required column: {col}")

    current_app.logger.info(f"[etl] Starting processing (columns: {headers})")

    url   = current_app.config.get('DGRAPH_URL')
    token = current_app.config.get('DGRAPH_API_TOKEN')
    
    # Handle case where Dgraph is not configured
    if not url or not token:
        current_app.logger.warning("[etl] Dgraph not configured - skipping canonical data fetch")
        data = {}
    else:
        gql = """
        query {
          products: queryProduct { title productID }
          ingredients: queryIngredients { title ingredientID }
          certifications: queryCertification { title certID }
          allergens: queryAllergen { title charID }
        }
        """
        try:
            current_app.logger.info("[etl] Fetching canonical products/ingredients from Dgraph…")
            current_app.logger.info(f"[etl] Dgraph URL: {url}")
            current_app.logger.info(f"[etl] GraphQL Query: {gql}")
            resp = requests.post(url, json={"query": gql}, headers={"Content-Type": "application/json", "Dg-Auth": token}, timeout=10)
            resp.raise_for_status()
            response_data = resp.json() if resp else {}
            if response_data is None:
                response_data = {}
            current_app.logger.info(f"[etl] Raw Dgraph response: {response_data}")
            data = response_data.get("data", {})
            current_app.logger.info(f"[etl] Canonical products: {len(data.get('products',[]))}, ingredients: {len(data.get('ingredients',[]))}, certifications: {len(data.get('certifications',[]))}, allergens: {len(data.get('allergens',[]))}")
        except Exception as e:
            current_app.logger.error(f"[etl] Could not fetch canonical: {e}")
            current_app.logger.error(f"[etl] Dgraph URL: {url}")
            current_app.logger.error(f"[etl] GraphQL Query: {gql}")
            data = {}

    prod_map   = {p['title']: p['productID']   for p in data.get("products", [])}
    ing_map    = {i['title']: i['ingredientID'] for i in data.get("ingredients", [])}
    cert_map   = {c['title']: c['certID'] for c in data.get("certifications", [])}
    allergen_map = {a['title']: a['charID'] for a in data.get("allergens", [])}
    
    prod_lower = {t.lower(): ext for t,ext in prod_map.items()}
    ing_lower  = {t.lower(): ext for t,ext in ing_map.items()}
    cert_lower = {t.lower(): ext for t,ext in cert_map.items()}
    allergen_lower = {t.lower(): ext for t,ext in allergen_map.items()}
    
    prod_names = list(prod_map.keys())
    ing_names  = list(ing_map.keys())
    cert_names = list(cert_map.keys())
    allergen_names = list(allergen_map.keys())

    def is_valid(v):
        return not is_empty_or_invalid(v)

    validation_errors = []
    valid_row_indices = []
    counter = 0

    # Process rows one by one to avoid memory issues
    for idx, row in enumerate(rows_generator, start=2):
        current_app.logger.info(f"\n[etl] Processing row {idx}…")
        
        try:
            biz     = get(row, 'businessName')
            country = get(row, 'country1')
            row_errors = []
            if not is_valid(biz):
                row_errors.append("Missing or empty businessName")
                current_app.logger.warning(f"[etl] Row {idx}: Missing or empty businessName, skipping row.")
            if not is_valid(country):
                row_errors.append("Missing or empty country1")
                current_app.logger.warning(f"[etl] Row {idx}: Missing or empty country1, skipping row.")
            if row_errors:
                validation_errors.append({'row': idx, 'error': "; ".join(row_errors)})
                continue

            # Only add valid rows to DB!
            valid_row_indices.append(idx)
            current_app.logger.info(f"[etl] Row {idx}: Creating Member for '{biz}', country: '{country}'")
            
            # Sanitize and validate all inputs
            biz_sanitized = sanitize_string(biz)
            email_sanitized = sanitize_string(get(row, 'contactEmail'))
            address_sanitized = sanitize_string(get(row, 'streetAddress1'))
            city_sanitized = sanitize_string(get(row, 'city1'))
            country_sanitized = sanitize_string(country)
            bio_sanitized = sanitize_string(get(row, 'companyBio'))
            
            # Validate business name
            is_valid_name, name_error = validate_business_name(biz_sanitized)
            if not is_valid_name:
                validation_errors.append({'row': idx, 'error': f"Business name validation failed: {name_error}"})
                continue
            
            # Validate email if provided
            is_valid_email, email_error = validate_email(email_sanitized)
            if not is_valid_email:
                validation_errors.append({'row': idx, 'error': f"Email validation failed: {email_error}"})
                continue
            
            # Determine member offerings based on data presence
            # Use the full mapping (custom + auto-detected) for offerings detection
            full_mapping = custom_mapping or {}
            
            # Add mappings for the specific service columns that exist in the CSV
            service_column_mappings = {
                'manufacturingServices': 'manufacturingServices',
                'logisticalServices': 'logisticalServices', 
                'labServices': 'laboratoryServices',
                'startupFriendlyServices': 'startupFriendlyServices',
                'suppliedPackaging': 'suppliedPackaging',
                'deliveredIn': 'deliveredIn',
                'designServices': 'designServices',
                'legalServices': 'legalServices',
                'marketingServices': 'marketingServices',
                'regulatoryServices': 'regulatoryServices',
                'consultingServices': 'consultingServices',
                'facilityDetails': 'facilityDetails',
                'suppliedEquipment': 'suppliedEquipment',
                'ingredients': 'ingredients',
                'products': 'products'  # Products also indicate manufacturing
            }
            
            # Add mappings for any service columns that exist in the headers
            for header in headers:
                if header in service_column_mappings and header not in full_mapping:
                    full_mapping[header] = {
                        'schema_field': service_column_mappings[header],
                        'confidence': 100,  # Direct match
                        'original_header': header
                    }
            
            member_offerings = determine_member_offerings(row, full_mapping)
            current_app.logger.info(f"[etl] Row {idx}: Member offerings detected: {[o['title'] for o in member_offerings]}")
            
            member = Member(
                name=biz_sanitized,
                contact_email=email_sanitized or None,
                street_address1=address_sanitized or None,
                city1=city_sanitized or None,
                country1=country_sanitized,
                company_bio=bio_sanitized or None,
                submission=submission
            )
            db.session.add(member)
            db.session.flush()
            
            # Store member offerings in the database for persistence
            member.member_offerings = member_offerings
            
            # Also store in session cache for backward compatibility
            if not hasattr(db.session, 'member_offerings_cache'):
                db.session.member_offerings_cache = {}
            db.session.member_offerings_cache[member.id] = member_offerings

            def handle(kind, cell):
                nonlocal counter
                kind_mapping = {
                    'product': 'Product',
                    'ingredient': 'Ingredient', 
                    'certification': 'Certification',
                    'allergen': 'Allergen'
                }
                kindstr = kind_mapping.get(kind, kind.title())
                current_app.logger.info(f"[etl] Row {idx}: Handling {kindstr}s for member '{biz}'…")
                if not is_valid(cell):
                    current_app.logger.info(f"[etl] Row {idx}: No {kindstr}s listed.")
                    return
                fragments = re.split(r'[;,]', str(cell))
                
                # Deduplication: Track processed items to avoid duplicates within the same row
                processed_items = set()
                
                for raw in fragments:
                    text = raw.strip()
                    if not is_valid(text):
                        current_app.logger.info(f"[etl] Row {idx}: Skipping blank/invalid {kindstr}.")
                        continue
                    
                    # Normalize and sanitize item name
                    text_normalized = normalize_offering_text(text)
                    text_sanitized = sanitize_string(text_normalized)
                    
                    # Check for duplicates within this row
                    if text_sanitized.lower() in processed_items:
                        current_app.logger.info(f"[etl] Row {idx}: Skipping duplicate {kindstr}: '{text_sanitized}'")
                        continue
                    
                    # Add to processed items set
                    processed_items.add(text_sanitized.lower())
                    ni = NewItem(name=text_sanitized, type=kind, member=member)
                    
                    # Get the appropriate mapping based on kind
                    if kind == 'product':
                        lower_map = prod_lower
                        pool = prod_names
                        ext_map = prod_map
                    elif kind == 'ingredient':
                        lower_map = ing_lower
                        pool = ing_names
                        ext_map = ing_map
                    elif kind == 'certification':
                        lower_map = cert_lower
                        pool = cert_names
                        ext_map = cert_map
                    elif kind == 'allergen':
                        lower_map = allergen_lower
                        pool = allergen_names
                        ext_map = allergen_map
                    else:
                        current_app.logger.warning(f"[etl] Row {idx}: Unknown kind '{kind}', skipping")
                        continue

                    # Exact match
                    ext_id = lower_map.get(text_sanitized.lower())
                    if ext_id:
                        ni.matched_canonical_id = ext_id
                        ni.score = 100.0
                        ni.resolved = True
                        current_app.logger.info(f"[etl] Row {idx}: '{text_sanitized}' ({kindstr}) exact matched existing canonical [{ext_id}]")
                    else:
                        # Enhanced fuzzy matching with penalty-based ranking
                        # Get all potential matches with raw scores
                        all_matches = process.extract(text_sanitized, pool, scorer=fuzz.token_set_ratio, processor=utils.default_process, limit=10)
                        
                        # Apply penalties to all matches and find the best one
                        penalized_matches = []
                        best_match = None
                        best_adjusted_score = 0.0
                        
                        for match_name, raw_score, _ in all_matches:
                            # Additional scoring algorithms for cross-validation (only for the best raw match)
                            if match_name == all_matches[0][0]:  # Only for the raw best match
                                best_ratio = process.extractOne(text_sanitized, [match_name], scorer=fuzz.ratio, processor=utils.default_process)
                                ratio_score = float(best_ratio[1]) if best_ratio else 0.0
                                
                                best_partial = process.extractOne(text_sanitized, [match_name], scorer=fuzz.partial_ratio, processor=utils.default_process)
                                partial_score = float(best_partial[1]) if best_partial else 0.0
                                
                                # Apply penalties including algorithm disagreement
                                adjusted_score = apply_match_penalties(text_sanitized, match_name, float(raw_score))
                                
                                # Additional penalty for algorithm disagreement
                                score_variance = max(abs(float(raw_score) - ratio_score), abs(float(raw_score) - partial_score))
                                if score_variance > ALGORITHM_DISAGREEMENT_THRESHOLD:
                                    adjusted_score -= ALGORITHM_DISAGREEMENT_PENALTY
                                
                                current_app.logger.info(f"[etl] Row {idx}: Raw best '{match_name}' raw score: {raw_score:.1f}%, adjusted score: {adjusted_score:.1f}%")
                            else:
                                # Apply penalties without algorithm disagreement check
                                adjusted_score = apply_match_penalties(text_sanitized, match_name, float(raw_score))
                                current_app.logger.info(f"[etl] Row {idx}: Alternative '{match_name}' raw score: {raw_score:.1f}%, adjusted score: {adjusted_score:.1f}%")
                            
                            # Ensure score doesn't go below 0
                            adjusted_score = max(0.0, adjusted_score)
                            
                            penalized_matches.append((match_name, adjusted_score))
                            
                            # Track the best adjusted score
                            if adjusted_score > best_adjusted_score:
                                best_adjusted_score = adjusted_score
                                best_match = match_name
                        
                        # Now use the best match after penalties
                        name0 = best_match
                        final_score = best_adjusted_score
                        
                        current_app.logger.info(f"[etl] Row {idx}: '{text_sanitized}' ({kindstr}) best match after penalties: '{name0}' (score {final_score:.1f}%)")
                        
                        # Use configurable thresholds
                        threshold = get_fuzzy_match_threshold()
                        auto_resolve_threshold = get_auto_resolve_threshold()
                        auto_reject_threshold = get_auto_reject_threshold()
                        
                        # Implement proper threshold logic as per requirements:
                        # Auto-accept if score ≥ 95% (configurable)
                        # Auto-reject if score < 50% (but still show as candidate "No match")
                        # Flag for review if 50% ≤ score < 95% (thresholds adjustable)
                        if final_score >= auto_resolve_threshold:
                            # High confidence - auto-resolve (penalties already applied)
                            ni.matched_canonical_id = ext_map[name0]
                            ni.score = final_score
                            ni.resolved = True
                            current_app.logger.info(f"[etl] Row {idx}: '{text_sanitized}' ({kindstr}) auto-resolved with '{name0}' (score {final_score:.1f}%)")
                        elif final_score >= auto_reject_threshold:
                            # Medium confidence - create review but mark as suggested
                            ni.matched_canonical_id = ext_map[name0]
                            ni.score = final_score
                            ni.resolved = False  # Don't auto-resolve, require review
                            current_app.logger.info(f"[etl] Row {idx}: '{text_sanitized}' ({kindstr}) suggested match '{name0}' (score {final_score:.1f}%) - requires review")
                            
                            # Create review for suggested match with alternatives
                            # Use the already calculated penalized matches as alternatives
                            alts = []
                            
                            # Use penalized_matches but exclude the best match and filter out low-confidence matches
                            for alt_name, alt_score in penalized_matches:
                                if alt_name != name0 and alt_score >= auto_reject_threshold:  # Skip the best match and low-confidence matches
                                    alt_ext_id = ext_map.get(alt_name)
                                    alts.append({"name": alt_name, "score": alt_score, "ext_id": alt_ext_id})
                                    # Stop when we have 3 alternatives
                                    if len(alts) >= 3:
                                        break
                            
                            mr = MatchReview(
                                new_item=ni, suggested_name=name0,
                                suggested_ext_id=ext_map[name0],
                                score=final_score, alternatives=alts, approved=None
                            )
                            db.session.add(mr)
                        else:
                            # Low confidence - auto-reject (no alternatives shown since all are below threshold)
                            alts = []
                            suggested_ext_id = ext_map.get(name0) if name0 else None
                                    
                            current_app.logger.info(
                                f"[etl] Row {idx}: '{text_sanitized}' ({kindstr}) auto-rejected (score {final_score:.1f}% < {auto_reject_threshold}%) - no good match found. Top guess: '{name0}'."
                            )
                            mr = MatchReview(
                                new_item=ni, suggested_name=name0 or text_sanitized,
                                suggested_ext_id=suggested_ext_id,
                                score=final_score, alternatives=alts, approved=False
                            )
                            # Set ignored flag on the NewItem instead
                            ni.ignored = True
                            db.session.add(mr)
                    db.session.add(ni)
                    counter += 1
                    if counter % BATCH_SIZE == 0:
                        db.session.flush()  # Use flush instead of commit for better performance
                        current_app.logger.info(f"[etl] processed {counter} items…")

            handle('product', get(row, 'products'))
            handle('ingredient', get(row, 'ingredients'))
            handle('certification', get(row, 'certifications'))
            handle('allergen', get(row, 'allergens'))
        except SQLAlchemyError as ex:
            err_msg = f"DB error: {ex}"
            current_app.logger.error(f"[etl] Row {idx}: {err_msg}")
            validation_errors.append({'row': idx, 'error': err_msg})
            continue

    current_app.logger.info(f"[etl] Finished processing → {counter} unique items; {len(validation_errors)} rows skipped")
    if validation_errors:
        for err in validation_errors:
            current_app.logger.warning(f"[etl] Validation Error Row {err['row']}: {err['error']}")
    else:
        current_app.logger.info(f"[etl] No validation errors.")
    return counter, validation_errors, valid_row_indices

def convert_excel_to_csv_suggestion(filename):
    """Provide helpful suggestion for converting problematic Excel files to CSV"""
    return (
        f"💡 **Excel File Conversion Tip**\n\n"
        f"The file '{filename}' cannot be processed. Here's how to fix it:\n\n"
        f"**Option 1: Re-save as Excel (.xlsx)**\n"
        f"1. Open the file in Microsoft Excel or Google Sheets\n"
        f"2. Go to File → Save As\n"
        f"3. Choose 'Excel Workbook (.xlsx)'\n"
        f"4. Save and try uploading again\n\n"
        f"**Option 2: Convert to CSV (Recommended)**\n"
        f"1. Open the file in Excel/Google Sheets\n"
        f"2. Go to File → Save As\n"
        f"3. Choose 'CSV (Comma delimited) (*.csv)'\n"
        f"4. Upload the CSV file instead\n\n"
        f"**Why this happens:**\n"
        f"• File corruption during download/transfer\n"
        f"• Saved in an unsupported Excel format\n"
        f"• File extension doesn't match actual format\n"
        f"• File was created by non-Excel software"
    )


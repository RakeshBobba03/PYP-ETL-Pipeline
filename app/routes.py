# app/routes.py

import os
import csv
import requests
import logging
import time
import json
import openpyxl
from datetime import datetime
from pathlib import Path
from flask import (
    Blueprint, render_template, request, redirect,
    url_for, current_app, send_from_directory, session, abort, jsonify, make_response
)
from urllib.parse import quote
from werkzeug.utils import secure_filename
from app import db
from app.models import NewItem, MatchReview, MemberSubmission, Member, Product, Ingredient
from app.etl import process_submission_file, map_headers_to_schema, validate_required_columns, normalize_data_sample, get_member_offerings_from_cache, open_csv_with_encoding_detection
from app.logging_utils import logging_manager
from app.error_utils import error_handler, ErrorCategory
from app.report_utils import report_generator

main_bp = Blueprint('main', __name__)

UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')  # Change to any folder you want
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

# Custom Jinja filters
@main_bp.app_template_filter('confidence_class')
def confidence_class_filter(confidence):
    """Convert confidence score to CSS class"""
    if confidence >= 90:
        return 'high'
    elif confidence >= 70:
        return 'medium'
    else:
        return 'low'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def is_safe_filename(filename):
    """Check if filename is safe (no path traversal)"""
    if not filename:
        return False
    # Normalize path and check if it's within upload folder
    try:
        file_path = Path(UPLOAD_FOLDER) / filename
        file_path.resolve().relative_to(Path(UPLOAD_FOLDER).resolve())
        return True
    except (ValueError, RuntimeError):
        return False

def dgraph_request_with_retry(url, json_data, headers, max_retries=3, base_delay=1, operation_id=None):
    """Make Dgraph request with exponential backoff retry and enhanced error handling"""
    for attempt in range(max_retries):
        try:
            # Track data usage for daily limit monitoring
            data_size = error_handler.estimate_data_size(json_data)
            error_handler.track_data_usage(data_size, "mutation")
            
            # Check daily limit before making request
            limit_exceeded, usage_gb, limit_gb = error_handler.check_daily_limit()
            if limit_exceeded:
                error_msg = f"Daily limit exceeded: {usage_gb:.2f}GB / {limit_gb}GB"
                current_app.logger.error(f"[dgraph] {error_msg}")
                raise Exception(error_msg)
            
            resp = requests.post(url, json=json_data, headers=headers, timeout=30)
            resp.raise_for_status()
            
            # Log successful mutation
            response_data = resp.json()
            if response_data is None:
                response_data = {}
            logging_manager.log_mutation(
                mutation_type="dgraph_request",
                payload=json_data,
                response={"status": "success", "data": response_data},
                dgraph_url=url,
                headers=headers,
                operation_id=operation_id
            )
            
            return resp
            
        except requests.exceptions.RequestException as e:
            # Handle error with categorization
            error_info = error_handler.handle_error(
                error=e,
                context=f"Dgraph request (attempt {attempt + 1}/{max_retries})",
                operation_id=operation_id,
                retry_count=attempt,
                max_retries=max_retries
            )
            
            if not error_info["should_retry"] or attempt == max_retries - 1:
                # Log final failure
                logging_manager.log_mutation(
                    mutation_type="dgraph_request",
                    payload=json_data,
                    response={"status": "error", "errors": [str(e)]},
                    dgraph_url=url,
                    headers=headers,
                    operation_id=operation_id
                )
                raise e
            
            delay = error_info["retry_delay"]
            current_app.logger.warning(f"[dgraph] Request failed (attempt {attempt + 1}/{max_retries}), retrying in {delay}s: {e}")
            time.sleep(delay)
    
    raise Exception(f"All {max_retries} retry attempts failed")

def is_semantically_valid_match(original_name, suggested_name, item_type):
    """
    Perform additional semantic validation to prevent incorrect matches.
    Returns True if the match is semantically valid, False otherwise.
    """
    if not original_name or not suggested_name:
        return False
    
    original_lower = original_name.lower().strip()
    suggested_lower = suggested_name.lower().strip()
    
    # Define category-specific keywords that should not be mixed
    category_keywords = {
        'vitamins': ['vitamin', 'vitamins', 'vit', 'ascorbic', 'thiamine', 'riboflavin', 'niacin', 'b12', 'b6', 'folate', 'biotin', 'pantothenic'],
        'amino_acids': ['amino', 'acid', 'protein', 'peptide', 'glutamine', 'arginine', 'lysine', 'methionine', 'tryptophan', 'tyrosine'],
        'minerals': ['calcium', 'iron', 'zinc', 'magnesium', 'selenium', 'copper', 'manganese', 'chromium', 'iodine', 'phosphorus'],
        'omega': ['omega', 'dha', 'epa', 'fatty', 'acid', 'fish', 'oil', 'flax', 'linseed'],
        'probiotics': ['probiotic', 'probiotics', 'lactobacillus', 'bifidobacterium', 'acidophilus', 'bacteria', 'culture'],
        'prebiotics': ['prebiotic', 'prebiotics', 'fiber', 'inulin', 'fructooligosaccharide', 'galactooligosaccharide'],
        'certifications': ['organic', 'certified', 'usda', 'canada', 'european', 'bio', 'eco', 'sustainable', 'fair trade'],
        'additives': ['additive', 'additives', 'preservative', 'stabilizer', 'emulsifier', 'thickener', 'colorant'],
        'adhesives': ['adhesive', 'adhesives', 'glue', 'bonding', 'sealant', 'cement', 'paste']
    }
    
    # Check for category mismatches
    for category, keywords in category_keywords.items():
        original_has_category = any(keyword in original_lower for keyword in keywords)
        suggested_has_category = any(keyword in suggested_lower for keyword in keywords)
        
        # If one has the category and the other doesn't, it's likely a mismatch
        if original_has_category != suggested_has_category:
            # Special case: allow some flexibility for similar categories
            if category == 'omega' and ('omega' in original_lower or 'omega' in suggested_lower):
                # Allow omega-3 to match omega-6, but not other categories
                continue
            elif category == 'probiotics' and category == 'prebiotics':
                # These are related but different - be more strict
                continue
            else:
                current_app.logger.info(f"[semantic_validation] Category mismatch: '{original_name}' ({category}) vs '{suggested_name}'")
                return False
    
    # Check for specific problematic patterns
    problematic_patterns = [
        # Vitamin vs Amino Acid mismatches
        ('vitamin', 'amino'),
        ('vitamin', 'protein'),
        ('vitamin', 'peptide'),
        # Additive vs Adhesive mismatches  
        ('additive', 'adhesive'),
        ('additive', 'glue'),
        ('additive', 'bonding'),
        # Probiotic vs Prebiotic mismatches
        ('probiotic', 'prebiotic'),
        ('bacteria', 'fiber'),
        ('culture', 'inulin'),
        # Mineral vs Vitamin mismatches
        ('calcium', 'vitamin'),
        ('iron', 'vitamin'),
        ('zinc', 'vitamin'),
    ]
    
    for pattern1, pattern2 in problematic_patterns:
        if pattern1 in original_lower and pattern2 in suggested_lower:
            current_app.logger.info(f"[semantic_validation] Problematic pattern: '{original_name}' ({pattern1}) vs '{suggested_name}' ({pattern2})")
            return False
        if pattern2 in original_lower and pattern1 in suggested_lower:
            current_app.logger.info(f"[semantic_validation] Problematic pattern: '{original_name}' ({pattern2}) vs '{suggested_name}' ({pattern1})")
            return False
    
    # Check for length difference (too different lengths might indicate different items)
    length_ratio = min(len(original_lower), len(suggested_lower)) / max(len(original_lower), len(suggested_lower))
    if length_ratio < 0.5:  # If one is less than half the length of the other
        current_app.logger.info(f"[semantic_validation] Length mismatch: '{original_name}' ({len(original_lower)}) vs '{suggested_name}' ({len(suggested_lower)})")
        return False
    
    # If we get here, the match passes semantic validation
    return True

@main_bp.route('/')
def index():
    current_app.logger.info("[routes] Redirecting to upload page")
    return redirect(url_for('main.upload_file'))

@main_bp.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        file = request.files.get('file')
        clear_previous = request.form.get('clear_previous') == 'on'
        current_app.logger.info(f"[upload] POST received. File: {file.filename if file else 'None'}, clear_previous: {clear_previous}")

        session.pop('etl_validation_errors', None)
        session.pop('etl_error_filename', None)
        session.pop('custom_mapping', None)
        session.pop('updated_mapping', None)
        session.pop('updated_validation', None)
        # Note: updated_sample_data is no longer stored in session

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            
            # Ensure upload directory exists
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
            
            save_path = os.path.join(UPLOAD_FOLDER, filename)
            file.save(save_path)
            current_app.logger.info(f"[upload] File uploaded and saved to: {save_path}")

            # Store file info in session for validation flow
            session['uploaded_file'] = filename
            session['file_path'] = save_path

            if clear_previous:
                current_app.logger.info("[upload] Clearing previous submissions and DB records…")
                MatchReview.query.delete()
                NewItem.query.delete()
                Member.query.delete()
                MemberSubmission.query.delete()
                db.session.commit()
                current_app.logger.info("[upload] Previous DB records cleared.")

            # Redirect to validation page instead of processing immediately
            return redirect(url_for('main.validate_headers'))

        current_app.logger.warning("[upload] No valid file selected or invalid file type.")
        # Redirect back to upload with error message
        return redirect(url_for('main.upload_file', status='error', message='Please select a valid .xlsx, .xls or .csv file.'))

    current_app.logger.info("[upload] GET received. Rendering upload.html")
    return render_template('upload.html')

@main_bp.route('/validate_headers')
def validate_headers():
    """Show header mapping and validation results"""
    filename = session.get('uploaded_file')
    file_path = session.get('file_path')
    
    if not filename or not file_path:
        return redirect(url_for('main.upload_file', status='error', message='No file uploaded. Please upload a file first.'))
    
    if not os.path.exists(file_path):
        return redirect(url_for('main.upload_file', status='error', message='Uploaded file not found. Please upload again.'))
    
    try:
        # Extract headers from file
        ext = filename.lower().rsplit('.', 1)[1]
        headers = []
        
        if ext == 'csv':
            f, encoding = open_csv_with_encoding_detection(file_path)
            try:
                reader = csv.reader(f)
                headers = next(reader, [])
                current_app.logger.info(f"[validate_headers] Successfully read CSV with encoding: {encoding}")
            finally:
                f.close()
        elif ext in ['xlsx', 'xls']:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [h if h else '' for h in header_row]
            wb.close()
        
        # Check if we have updated mapping from session
        updated_mapping = session.get('updated_mapping')
        updated_validation = session.get('updated_validation')
        
        if updated_mapping and updated_validation:
            # Use updated data from session
            mapping = updated_mapping
            validation = updated_validation
            # Regenerate sample data to avoid storing large data in session
            sample_data = normalize_data_sample(file_path, headers, mapping, sample_size=10)
            unmapped = [h for h in headers if h not in mapping]
            
            # Don't clear session data yet - keep it for potential refreshes
            # Only clear when user proceeds to next step
        else:
            # Use automatic mapping
            mapping, unmapped = map_headers_to_schema(headers)
            validation = validate_required_columns(headers, mapping)
            sample_data = normalize_data_sample(file_path, headers, mapping, sample_size=10)
        
        # Get schema fields and offerings mapping for template
        from app.etl import get_schema_field_mapping, get_member_offerings_mapping
        schema_fields = get_schema_field_mapping()
        offerings_mapping = get_member_offerings_mapping()
        
        return render_template(
            'validate_headers.html',
            filename=filename,
            headers=headers,
            mapping=mapping,
            unmapped=unmapped,
            validation=validation,
            sample_data=sample_data,
            schema_fields=schema_fields,
            offerings_mapping=offerings_mapping
        )
        
    except Exception as e:
        current_app.logger.error(f"Error validating headers: {e}")
        error_msg = str(e)
        
        # Log the error and redirect back to upload
        current_app.logger.error(f"Error validating file: {error_msg}")
        
        return redirect(url_for('main.upload_file', status='error', message=quote(f'Error validating file: {error_msg}')))

@main_bp.route('/update_mapping', methods=['POST'])
def update_mapping():
    """Update header mapping based on user input"""
    current_app.logger.info(f"[update_mapping] POST received. Form data: {request.form}")
    current_app.logger.info(f"[update_mapping] Is JSON: {request.is_json}")
    
    filename = session.get('uploaded_file')
    file_path = session.get('file_path')
    
    if not filename or not file_path:
        current_app.logger.error("[update_mapping] No file uploaded")
        return jsonify({'error': 'No file uploaded'}), 400
    
    try:
        # Handle both form data and JSON data
        if request.is_json:
            data = request.get_json()
            custom_mapping = data.get('mapping', {})
        else:
            # Handle form data
            mapping_json = request.form.get('mapping', '{}')
            try:
                custom_mapping = json.loads(mapping_json)
            except json.JSONDecodeError:
                if request.is_json:
                    return jsonify({'error': 'Invalid mapping data format'}), 400
                else:
                    return redirect(url_for('main.validate_headers', status='error', message='Invalid mapping data format'))
        
        # Store custom mapping in session
        session['custom_mapping'] = custom_mapping
        
        # Re-validate with custom mapping
        ext = filename.lower().rsplit('.', 1)[1]
        headers = []
        
        if ext == 'csv':
            f, encoding = open_csv_with_encoding_detection(file_path)
            try:
                reader = csv.reader(f)
                headers = next(reader, [])
                current_app.logger.info(f"[update_mapping] Successfully read CSV with encoding: {encoding}")
            finally:
                f.close()
        elif ext in ['xlsx', 'xls']:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [h if h else '' for h in header_row]
            wb.close()
        
        # Apply custom mapping
        mapping = {}
        for header in headers:
            if header in custom_mapping:
                mapping[header] = {
                    'schema_field': custom_mapping[header],
                    'confidence': 100,  # User-defined mapping
                    'original_header': header
                }
        
        # Validate required columns
        validation = validate_required_columns(headers, mapping)
        
        # Generate updated data sample
        sample_data = normalize_data_sample(file_path, headers, mapping, sample_size=10)
        
        # Store the updated data in session for the next page load
        # Only store essential data to avoid cookie size limits
        session['updated_mapping'] = mapping
        session['updated_validation'] = validation
        # Don't store sample_data in session - it's too large and will be regenerated
        
        # Return appropriate response based on request type
        if request.is_json:
            return jsonify({
                'success': True,
                'message': 'Mapping updated successfully',
                'validation': {
                    'is_valid': validation['is_valid'],
                    'missing_fields': validation['missing_fields'],
                    'total_headers': validation['total_headers'],
                    'mapped_headers': validation['mapped_headers']
                }
            })
        else:
            # Redirect back to validation page with updated data
            return redirect(url_for('main.validate_headers'))
        
    except Exception as e:
        current_app.logger.error(f"Error updating mapping: {e}")
        error_msg = str(e)
        
        # Provide specific guidance for Excel file errors
        if "Bad offset for central directory" in error_msg or "BadZipFile" in error_msg:
            if request.is_json:
                return jsonify({
                    'error': 'Excel file is corrupted. Please re-save the file in Excel or convert to CSV format.'
                }), 500
            else:
                return redirect(url_for('main.validate_headers', status='error', message='Excel file is corrupted. Please re-save the file in Excel or convert to CSV format.'))
        else:
            if request.is_json:
                return jsonify({'error': str(e)}), 500
            else:
                return redirect(url_for('main.validate_headers', status='error', message=str(e)))

@main_bp.route('/process_validated_file', methods=['POST'])
def process_validated_file():
    """Process the file after validation and mapping confirmation"""
    filename = session.get('uploaded_file')
    file_path = session.get('file_path')
    
    if not filename or not file_path:
        return redirect(url_for('main.upload_file', status='error', message='No file uploaded. Please upload a file first.'))
    
    try:
        # Get custom mapping from session
        custom_mapping = session.get('custom_mapping', {})
        current_app.logger.info(f"[process_validated_file] Starting ETL process for validated file: {filename}")
        current_app.logger.info(f"[process_validated_file] Custom mapping: {custom_mapping}")
        count, val_errors, valid_row_indices = process_submission_file(filename, custom_mapping=custom_mapping)
        current_app.logger.info(f"[process_validated_file] ETL finished for {filename}: {count} items, {len(val_errors)} validation errors")
        
        # Clear session data
        session.pop('uploaded_file', None)
        session.pop('file_path', None)
        session.pop('custom_mapping', None)
        session.pop('updated_mapping', None)
        session.pop('updated_validation', None)
        # Note: updated_sample_data is no longer stored in session
        
        # Handle results similar to original upload flow
        if count == 0:
            error_filename = f"{filename.rsplit('.',1)[0]}_errors.csv"
            error_path = os.path.join(UPLOAD_FOLDER, error_filename)
            current_app.logger.warning(f"[process_validated_file] All rows invalid for {filename}. Writing errors to: {error_path}")
            with open(error_path, 'w', newline='', encoding='utf-8') as ef:
                writer = csv.writer(ef)
                writer.writerow(['Row','Error'])
                for err in val_errors:
                    writer.writerow([err['row'], err['error']])
            return render_template(
                'etl_errors.html',
                submission=filename,
                errors=val_errors,
                error_filename=error_filename
            )

        # Store validation errors in session for banner/download on review page
        if val_errors:
            session['etl_validation_errors'] = val_errors
            session['etl_error_filename'] = f"{filename.rsplit('.',1)[0]}_errors.csv"
            error_path = os.path.join(UPLOAD_FOLDER, session['etl_error_filename'])
            current_app.logger.warning(f"[process_validated_file] Some rows were skipped. Writing ETL error report to: {error_path}")
            with open(error_path, 'w', newline='', encoding='utf-8') as ef:
                writer = csv.writer(ef)
                writer.writerow(['Row','Error'])
                for err in val_errors:
                    writer.writerow([err['row'], err['error']])
        current_app.logger.info(f"[process_validated_file] Successfully processed {count} valid items from {filename}")
        return redirect(url_for('main.review_list', status='error', message='Review item not found or already handled.'))

    except Exception as e:
        current_app.logger.error(f"[process_validated_file][error] {e}")
        error_msg = str(e)
        
        # Log the error and redirect back to validation
        current_app.logger.error(f"Processing failed: {e}")
        return redirect(url_for('main.validate_headers', status='error', message=quote(f'Processing failed: {str(e)}')))

@main_bp.route('/download_etl_errors')
def download_etl_errors():
    errors = session.get('etl_validation_errors')
    filename = session.get('etl_error_filename', 'ETL_Errors.csv')
    current_app.logger.info(f"[download_etl_errors] Download request. errors present: {bool(errors)}, filename: {filename}")
    if not errors or not filename:
        return redirect(url_for('main.review_list', status='error', message='Review item not found or already handled.'))
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)

@main_bp.route('/errors/<filename>')
def download_errors(filename):
    current_app.logger.info(f"[download_errors] Downloading error file: {filename}")
    
    # Security check: prevent path traversal
    if not is_safe_filename(filename):
        current_app.logger.warning(f"[download_errors] Attempted path traversal: {filename}")
        abort(400, description="Invalid filename")
    
    # Additional check: ensure file exists and is in allowed directory
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    if not os.path.exists(file_path) or not os.path.isfile(file_path):
        abort(404, description="File not found")
    
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)

@main_bp.route('/reviews')
def review_list():
    val_errors = session.get('etl_validation_errors')
    error_filename = session.get('etl_error_filename')
    current_app.logger.info("[review_list] Checking for pending reviews…")
    pending = MatchReview.query.join(NewItem) \
        .filter(MatchReview.approved.is_(None), NewItem.ignored.is_(False)) \
        .all()
    if pending:
        current_app.logger.info(f"[review_list] {len(pending)} pending reviews found. Rendering reviews.html")
        return render_template('reviews.html', pending_reviews=pending, val_errors=val_errors, error_filename=error_filename)

    new_items_to_add = NewItem.query.filter_by(resolved=False).all()
    
    # Fix: Properly categorize approved items as new vs matched
    # New items: approved but NOT resolved (user chose "Create New")
    # Matched items: approved AND resolved (user chose existing match)
    new_items_approved = NewItem.query.join(MatchReview).filter(
        NewItem.resolved == False,  # Fixed: Look for unresolved items (Create New)
        MatchReview.approved == True
    ).all()
    
    # Get all items that were resolved (either auto-resolved or manually matched)
    # This includes items that were linked to existing canonical data
    matched_items = NewItem.query.filter(
        NewItem.resolved == True,   # Look for resolved items (canonical matches)
        NewItem.ignored == False    # Exclude ignored items
    ).all()
    
    # Fetch all canonical data in bulk for caching
    canonical_titles = {'product': {}, 'ingredient': {}, 'certification': {}}
    
    try:
        url = current_app.config.get('DGRAPH_URL')
        token = current_app.config.get('DGRAPH_API_TOKEN')
        
        if url and token:
            headers = {"Content-Type": "application/json", "Dg-Auth": token}
            
            # Fetch all products
            product_query = """
            query {
                queryProduct {
                    productID
                    title
                }
            }
            """
            product_response = requests.post(url, json={"query": product_query}, headers=headers, timeout=10)
            if product_response.status_code == 200:
                product_data = product_response.json() if product_response else {}
                if product_data is None:
                    product_data = {}
                if product_data and 'data' in product_data:
                    for product in product_data['data'].get('queryProduct', []):
                        canonical_titles['product'][product['productID']] = product['title']
                    current_app.logger.info(f"[review_list] Fetched {len(canonical_titles['product'])} products from Dgraph")
            
            # Fetch all ingredients
            ingredient_query = """
            query {
                queryIngredients {
                    ingredientID
                    title
                }
            }
            """
            ingredient_response = requests.post(url, json={"query": ingredient_query}, headers=headers, timeout=10)
            if ingredient_response.status_code == 200:
                ingredient_data = ingredient_response.json() if ingredient_response else {}
                if ingredient_data is None:
                    ingredient_data = {}
                if ingredient_data and 'data' in ingredient_data:
                    for ingredient in ingredient_data['data'].get('queryIngredients', []):
                        canonical_titles['ingredient'][ingredient['ingredientID']] = ingredient['title']
                    current_app.logger.info(f"[review_list] Fetched {len(canonical_titles['ingredient'])} ingredients from Dgraph")
            
            # Fetch all certifications
            certification_query = """
            query {
                queryCertification {
                    certID
                    title
                }
            }
            """
            certification_response = requests.post(url, json={"query": certification_query}, headers=headers, timeout=10)
            if certification_response.status_code == 200:
                certification_data = certification_response.json() if certification_response else {}
                if certification_data is None:
                    certification_data = {}
                if certification_data and 'data' in certification_data:
                    for certification in certification_data['data'].get('queryCertification', []):
                        canonical_titles['certification'][certification['certID']] = certification['title']
                    current_app.logger.info(f"[review_list] Fetched {len(canonical_titles['certification'])} certifications from Dgraph")
            
            current_app.logger.info(f"[review_list] Total canonical data: {len(canonical_titles['product'])} products, {len(canonical_titles['ingredient'])} ingredients, {len(canonical_titles['certification'])} certifications")
            
        else:
            current_app.logger.warning(f"[review_list] Dgraph not configured - URL: {url}, Token: {'present' if token else 'missing'}")
            
    except Exception as e:
        current_app.logger.error(f"[review_list] Error fetching canonical data: {e}")
        canonical_titles = {'product': {}, 'ingredient': {}, 'certification': {}}
    
    # For each matched item, get the canonical item name from cache
    for item in matched_items:
        if item.matched_canonical_id and not hasattr(item, 'canonical_name'):
            canonical_name = "Unknown"
            
            # Try to get canonical name from cache
            if item.type == 'product' and item.matched_canonical_id in canonical_titles['product']:
                canonical_name = canonical_titles['product'][item.matched_canonical_id]
                current_app.logger.info(f"[review_list] Found product name for {item.matched_canonical_id}: {canonical_name}")
            elif item.type == 'ingredient' and item.matched_canonical_id in canonical_titles['ingredient']:
                canonical_name = canonical_titles['ingredient'][item.matched_canonical_id]
                current_app.logger.info(f"[review_list] Found ingredient name for {item.matched_canonical_id}: {canonical_name}")
            elif item.type == 'certification' and item.matched_canonical_id in canonical_titles['certification']:
                canonical_name = canonical_titles['certification'][item.matched_canonical_id]
                current_app.logger.info(f"[review_list] Found certification name for {item.matched_canonical_id}: {canonical_name}")
            else:
                # Fallback to suggested name if available
                if hasattr(item, 'review') and item.review and item.review.suggested_name:
                    canonical_name = item.review.suggested_name
                    current_app.logger.info(f"[review_list] Using suggested_name as fallback for {item.matched_canonical_id}: {canonical_name}")
                else:
                    current_app.logger.warning(f"[review_list] No canonical name found for {item.type} {item.matched_canonical_id}")
            
            # Store the canonical name as a dynamic attribute
            item.canonical_name = canonical_name
    
    # Get all companies that were created during ETL processing
    # This includes companies with auto-resolved items (no manual review needed)
    all_etl_companies = Member.query.all()
    
    # Get companies that have items requiring manual review (for the summary)
    companies_with_reviewed_items = set()
    for item in new_items_approved:
        companies_with_reviewed_items.add(item.member.name)
    
    current_app.logger.info(f"[review_list] No pending reviews. New items to add: {len(new_items_to_add)} | New items approved: {len(new_items_approved)} | Matched items: {len(matched_items)} | Total ETL companies: {len(all_etl_companies)}")
    return render_template(
        'reviews_done.html',
        new_items_to_add=new_items_to_add,
        new_items_approved=new_items_approved,
        matched_items=matched_items,
        all_etl_companies=all_etl_companies,
        companies_with_reviewed_items=companies_with_reviewed_items,
        val_errors=val_errors,
        error_filename=error_filename
    )

@main_bp.route('/reviews/handle_review/<int:item_id>', methods=['POST'])
def handle_review(item_id):
    current_app.logger.info(f"[handle_review] POST for review item_id={item_id}")
    review = MatchReview.query.filter_by(new_item_id=item_id, approved=None).first()
    if not review:
        current_app.logger.warning(f"[handle_review] Review item {item_id} not found or already handled.")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Review item not found or already handled.'})
        return redirect(url_for('main.review_list', status='error', message='Review item not found or already handled.'))

    # Handle multiple canonical choices
    canonical_choices = request.form.getlist('canonical_choices')
    single_choice = request.form.get('choice')
    
    current_app.logger.info(f"[handle_review] Canonical choices: {canonical_choices}, Single choice: {single_choice}")

    try:
        if canonical_choices:
            # User selected one or more canonical matches
            review.approved = True
            review.new_item.resolved = True
            
            # Store multiple canonical IDs as JSON in a new field
            # For now, we'll use the first choice as the primary match
            # and store the rest in the alternatives field
            primary_choice = canonical_choices[0]
            review.new_item.matched_canonical_id = primary_choice
            
            # Store all selected choices in alternatives for future reference
            all_choices = [{"ext_id": choice, "selected": True} for choice in canonical_choices]
            review.alternatives = all_choices
            
            message = f"Approved '{review.new_item.name}' matched to {len(canonical_choices)} canonical item(s)."
            current_app.logger.info(f"[handle_review] Approved '{review.new_item.name}' matched to {len(canonical_choices)} canonical items: {canonical_choices}")
            
        elif single_choice == '__new__':
            # User chose to create as new item
            review.approved = True
            review.new_item.resolved = False  # Mark as unresolved so it gets created as new
            # Don't set matched_canonical_id since we want it as a new item
            message = f"Approved '{review.new_item.name}' as new {review.new_item.type}."
            current_app.logger.info(f"[handle_review] Approved '{review.new_item.name}' as new {review.new_item.type}")
            
        elif single_choice and single_choice != '__new__':
            # User chose a single canonical match (backward compatibility)
            review.approved = True
            review.new_item.resolved = True
            review.new_item.matched_canonical_id = single_choice
            message = f"Approved '{review.new_item.name}' matched to canonical data."
            current_app.logger.info(f"[handle_review] Approved '{review.new_item.name}' matched to canonical ID: {single_choice}")
            
        else:
            # No choice made - treat as ignored
            review.approved = False
            review.new_item.ignored = True
            message = f"Ignored '{review.new_item.name}'."
            current_app.logger.info(f"[handle_review] Ignored review for item '{review.new_item.name}'")

        db.session.commit()
        
        # Return JSON for AJAX requests, redirect for regular requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': message})
        
        # Redirect back to review list
        return redirect(url_for('main.review_list', status='success', message=message))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[handle_review] Error processing review: {e}")
        error_message = f"Error processing review: {str(e)}"
        
        # Return JSON for AJAX requests, redirect for regular requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': error_message})
        
        return redirect(url_for('main.review_list', status='error', message='Review item not found or already handled.'))

@main_bp.route('/reviews/ignore_review_item/<int:item_id>', methods=['POST'])
def ignore_review_item(item_id):
    current_app.logger.info(f"[ignore_review_item] POST for item_id={item_id}")
    review = MatchReview.query.filter_by(new_item_id=item_id, approved=None).first()
    if not review:
        current_app.logger.warning(f"[ignore_review_item] Review item {item_id} not found or already handled.")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Review item not found or already handled.'})
        return redirect(url_for('main.review_list', status='error', message='Review item not found or already handled.'))

    try:
        review.approved = False
        review.new_item.ignored = True
        db.session.commit()
        
        message = f"Ignored '{review.new_item.name}'."
        current_app.logger.info(f"[ignore_review_item] Ignored review for item '{review.new_item.name}'")
        
        # Return JSON for AJAX requests, redirect for regular requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': message})
        
        return redirect(url_for('main.review_list', status='success', message=message))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[ignore_review_item] Error ignoring review: {e}")
        error_message = f"Error ignoring review: {str(e)}"
        
        # Return JSON for AJAX requests, redirect for regular requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': error_message})
        
        return redirect(url_for('main.review_list', status='error', message='Review item not found or already handled.'))

@main_bp.route('/reviews/batch_save_decisions', methods=['POST'])
def batch_save_decisions():
    reviews = MatchReview.query.join(NewItem) \
        .filter(MatchReview.approved.is_(None), NewItem.ignored.is_(False)) \
        .all()
    current_app.logger.info(f"[batch_save_decisions] Saving batch review decisions for {len(reviews)} items as NEW items")
    
    for review in reviews:
        # For batch save, we'll approve all items as "new" since they need review
        # This means they'll be created as new products/ingredients in Dgraph
        review.approved = True
        review.new_item.resolved = False  # Mark as unresolved so they get created as new
        # Don't set matched_canonical_id since we want them as new items
        current_app.logger.info(f"[batch_save_decisions] Marking '{review.new_item.name}' as new {review.new_item.type}")
        db.session.add(review)
        db.session.add(review.new_item)

    db.session.commit()
    current_app.logger.info(f"[batch_save_decisions] Batch review decisions saved for {len(reviews)} items as NEW items.")
    return redirect(url_for('main.review_list', status='success', message=quote(f'All {len(reviews)} items approved as NEW products/ingredients and ready to push to Dgraph.')))

@main_bp.route('/reviews/batch_approve_high_confidence', methods=['POST'])
def batch_approve_high_confidence():
    """Auto-approve items with high confidence matches (90% to 95%) with additional semantic validation"""
    from app.etl import get_fuzzy_match_threshold, get_auto_resolve_threshold
    
    # Use stricter thresholds for high confidence approval
    strict_fuzzy_threshold = 90.0  # Increased from 80% to 90%
    auto_resolve_threshold = get_auto_resolve_threshold()  # 95%
    
    # Get all potential high confidence reviews
    potential_reviews = MatchReview.query.join(NewItem) \
        .filter(
            MatchReview.approved.is_(None), 
            NewItem.ignored.is_(False),
            MatchReview.score >= strict_fuzzy_threshold,
            MatchReview.score < auto_resolve_threshold,  # Less than 95% (not auto-resolved)
            MatchReview.suggested_ext_id.isnot(None)
        ).all()
    
    current_app.logger.info(f"[batch_approve_high_confidence] Found {len(potential_reviews)} potential high confidence items")
    
    # Apply additional semantic validation
    approved_reviews = []
    rejected_reviews = []
    
    for review in potential_reviews:
        if is_semantically_valid_match(review.new_item.name, review.suggested_name, review.new_item.type):
            approved_reviews.append(review)
        else:
            rejected_reviews.append(review)
            current_app.logger.warning(f"[batch_approve_high_confidence] Rejected semantic mismatch: '{review.new_item.name}' -> '{review.suggested_name}' (score: {review.score:.1f}%)")
    
    # Auto-approve only the semantically valid matches
    approved_count = 0
    for review in approved_reviews:
        review.approved = True
        review.new_item.resolved = True
        review.new_item.matched_canonical_id = review.suggested_ext_id
        db.session.add(review)
        db.session.add(review.new_item)
        approved_count += 1

    db.session.commit()
    current_app.logger.info(f"[batch_approve_high_confidence] Auto-approved {approved_count} high confidence items (rejected {len(rejected_reviews)} semantic mismatches).")
    message = f'Auto-approved {approved_count} high confidence items ({strict_fuzzy_threshold}% to {auto_resolve_threshold}% match with semantic validation). {len(rejected_reviews)} items rejected due to semantic mismatches.'
    return redirect(url_for('main.review_list', status='success', message=quote(message)))

@main_bp.route('/reviews/batch_ignore_all', methods=['POST'])
def batch_ignore_all():
    pending = MatchReview.query.join(NewItem) \
        .filter(MatchReview.approved.is_(None), NewItem.ignored.is_(False)) \
        .all()
    current_app.logger.info(f"[batch_ignore_all] Ignoring all ({len(pending)}) pending review items")
    for review in pending:
        review.approved = False
        review.new_item.ignored = True
        db.session.add(review)
        db.session.add(review.new_item)

    db.session.commit()
    current_app.logger.info("[batch_ignore_all] All pending review items ignored.")
    return redirect(url_for('main.review_list', status='warning', message='All pending items have been ignored.'))

@main_bp.route('/reviews/preview_mutations')
def preview_mutations():
    """Preview the mutations that will be sent to Dgraph before actual submission"""
    submission = MemberSubmission.query.order_by(MemberSubmission.id.desc()).first()
    if not submission:
        current_app.logger.warning("[preview_mutations] No submission found to preview.")
        return redirect(url_for('main.upload_file', status='warning', message='No submission found to preview.'))

    url = current_app.config.get('DGRAPH_URL')
    token = current_app.config.get('DGRAPH_API_TOKEN')
    
    # Check if Dgraph is configured
    if not url or not token:
        current_app.logger.error("[preview_mutations] Dgraph not configured - cannot preview mutations")
        return redirect(url_for('main.review_list', status='error', message='Dgraph is not configured. Please set DGRAPH_URL and DGRAPH_API_TOKEN environment variables.'))
    
    headers = {"Content-Type": "application/json", "Dg-Auth": token}
    
    current_app.logger.info(f"[preview_mutations] Generating mutation preview for submission: {submission.name}")
    members = Member.query.filter_by(submission_id=submission.id).all()
    current_app.logger.info(f"[preview_mutations] Found {len(members)} member record(s) to preview")

    # Generate preview mutations (limit to first 3-5 members for preview)
    preview_members = members[:5]  # Show first 5 members as preview
    preview_mutations = []
    
    # Load valid countries from schema
    def load_valid_countries_from_schema():
        try:
            with open('listallcountries.json', 'r') as f:
                data = json.load(f)
                countries = data.get('data', {}).get('queryMemberCountry', [])
                return {country['title']: country['countryID'] for country in countries}
        except Exception as e:
            current_app.logger.warning(f"[preview_mutations] Could not load valid countries from schema: {e}")
            return {}
    
    valid_countries_schema = load_valid_countries_from_schema()
    
    def lookup_ref_preview(ref_type_query, var_name, title, id_field):
        """Lookup for preview - return actual ID from schema for countries"""
        if ref_type_query == "queryMemberCountry" and title in valid_countries_schema:
            return {id_field: valid_countries_schema[title]}
        else:
            return {id_field: f"<{id_field}_placeholder>"}
    
    for m in preview_members:
        biz = m.name or "(Unknown)"
        try:
            # Country lookup for preview
            if not m.country1 or m.country1 not in valid_countries_schema:
                continue
                
            country_ref = lookup_ref_preview("queryMemberCountry", "country", m.country1, "countryID")
            
            # Get all products and ingredients for this member
            all_products = [ni for ni in m.new_items if ni.type=="product" and not ni.ignored]
            all_ingredients = [ni for ni in m.new_items if ni.type=="ingredient" and not ni.ignored]
            
            # Separate resolved vs unresolved items
            resolved_products = [ni for ni in all_products if ni.resolved and ni.matched_canonical_id]
            unresolved_products = [ni for ni in all_products if not ni.resolved]
            
            resolved_ingredients = [ni for ni in all_ingredients if ni.resolved and ni.matched_canonical_id]
            unresolved_ingredients = [ni for ni in all_ingredients if not ni.resolved]
            
            # Collect product IDs (existing + resolved + new)
            existing_product_ids = []
            new_product_names = []
            
            # Add resolved product IDs
            for ni in resolved_products:
                if hasattr(ni, 'review') and ni.review and ni.review.alternatives:
                    # Multiple selections
                    for alt in ni.review.alternatives:
                        if isinstance(alt, dict) and alt.get('selected'):
                            existing_product_ids.append(alt.get('ext_id'))
                else:
                    # Single selection
                    if ni.matched_canonical_id:
                        existing_product_ids.append(ni.matched_canonical_id)
            
            # Check unresolved products
            for ni in unresolved_products:
                new_product_names.append(ni.name)
            
            # Same logic for ingredients
            existing_ingredient_ids = []
            new_ingredient_names = []
            
            for ni in resolved_ingredients:
                if hasattr(ni, 'review') and ni.review and ni.review.alternatives:
                    for alt in ni.review.alternatives:
                        if isinstance(alt, dict) and alt.get('selected'):
                            existing_ingredient_ids.append(alt.get('ext_id'))
                else:
                    if ni.matched_canonical_id:
                        existing_ingredient_ids.append(ni.matched_canonical_id)
            
            for ni in unresolved_ingredients:
                new_ingredient_names.append(ni.name)
            
            # Build member input for preview
            member_input = {
                "businessName": biz,
                "country1": country_ref,
                "streetAddress1": m.street_address1 if (m.street_address1 and m.street_address1.strip()) else "Not provided",
            }
            
            if m.contact_email and m.contact_email.strip():
                member_input["contactEmail"] = m.contact_email
            if m.city1 and m.city1.strip():
                member_input["city1"] = m.city1
            if m.company_bio and m.company_bio.strip():
                member_input["companyBio"] = m.company_bio
            
            # Add products and ingredients
            all_product_ids = existing_product_ids + [f"<new_product_{name}>" for name in new_product_names]
            all_ingredient_ids = existing_ingredient_ids + [f"<new_ingredient_{name}>" for name in new_ingredient_names]
            
            if all_product_ids:
                member_input["products"] = [{"productID": pid} for pid in all_product_ids]
            if all_ingredient_ids:
                member_input["ingredients"] = [{"ingredientID": iid} for iid in all_ingredient_ids]
            
            # Add member offerings
            member_offerings = get_member_offerings_from_cache(m.id)
            current_app.logger.info(f"[preview_mutations] Member '{biz}' (ID: {m.id}) offerings: {member_offerings}")
            
            # If no offerings found, try to get them from session cache or re-process
            if not member_offerings:
                # Try to get from session cache directly
                if hasattr(db.session, 'member_offerings_cache') and db.session.member_offerings_cache:
                    member_offerings = db.session.member_offerings_cache.get(m.id, [])
                    current_app.logger.info(f"[preview_mutations] Retrieved offerings from session cache for member '{biz}': {member_offerings}")
            
            if member_offerings:
                offering_refs = []
                for offering in member_offerings:
                    if isinstance(offering, dict) and offering is not None and 'uid' in offering:
                        offering_refs.append({"offeringID": offering['uid']})
                if offering_refs:
                    member_input["memberOfferings"] = offering_refs
                    current_app.logger.info(f"[preview_mutations] Added {len(offering_refs)} offerings to member '{biz}'")
            else:
                current_app.logger.warning(f"[preview_mutations] No offerings found for member '{biz}' (ID: {m.id})")
            
            # Create the mutation structure
            mutation_preview = {
                "member": member_input,
                "new_products": new_product_names,
                "new_ingredients": new_ingredient_names,
                "existing_products": len(existing_product_ids),
                "existing_ingredients": len(existing_ingredient_ids),
                "member_offerings": len(member_offerings) if member_offerings else 0
            }
            
            preview_mutations.append(mutation_preview)
            
        except Exception as ex:
            current_app.logger.warning(f"[preview_mutations] Error generating preview for '{biz}': {ex}")
            continue
    
    return render_template(
        'mutation_preview.html',
        submission=submission,
        preview_mutations=preview_mutations,
        total_members=len(members),
        preview_count=len(preview_mutations)
    )

@main_bp.route('/reviews/push', methods=['POST'])
def push_to_dgraph():
    submission = MemberSubmission.query.order_by(MemberSubmission.id.desc()).first()
    if not submission:
        current_app.logger.warning("[push_to_dgraph] No submission found to push.")
        return redirect(url_for('main.upload_file', status='warning', message='No submission found to push.'))

    url   = current_app.config.get('DGRAPH_URL')
    token = current_app.config.get('DGRAPH_API_TOKEN')
    
    # Check if Dgraph is configured
    if not url or not token:
        current_app.logger.error("[push] Dgraph not configured - cannot push data")
        return redirect(url_for('main.review_list', status='error', message='Dgraph is not configured. Please set DGRAPH_URL and DGRAPH_API_TOKEN environment variables.'))
    
    headers = {"Content-Type": "application/json", "Dg-Auth": token}
    
    # Generate operation ID for this push
    operation_id = logging_manager._generate_log_id("push_to_dgraph")
    
    # Log push start
    logging_manager.log_event("push_start", {
        "submission_name": submission.name,
        "submission_id": submission.id,
        "dgraph_url": url,
        "operation_id": operation_id
    }, operation_id)
    
    # Test Dgraph connectivity first
    try:
        test_query = {"query": "query { __schema { types { name } } }"}
        test_resp = requests.post(url, json=test_query, headers=headers, timeout=5)
        test_resp.raise_for_status()
        current_app.logger.info("[push] Dgraph connectivity test successful")
    except Exception as e:
        current_app.logger.error(f"[push] Dgraph connectivity test failed: {e}")
        return redirect(url_for('main.review_list', status='error', message=quote(f'Dgraph is not accessible: {str(e)}. Please check your Dgraph instance and daily limits.')))

    current_app.logger.info(f"[push] Starting push for submission: {submission.name}")
    members = Member.query.filter_by(submission_id=submission.id).all()
    current_app.logger.info(f"[push] Found {len(members)} member record(s) to process")

    results = {"members": [], "products": [], "ingredients": [], "errors": []}
    

    
    # Load valid countries from schema (listallcountries.json represents the schema)
    def load_valid_countries_from_schema():
        """Load all valid countries from the schema definition"""
        try:
            with open('listallcountries.json', 'r') as f:
                data = json.load(f)
                countries = data.get('data', {}).get('queryMemberCountry', [])
                return {country['title']: country['countryID'] for country in countries}
        except Exception as e:
            current_app.logger.warning(f"[push] Could not load valid countries from schema: {e}")
            return {}
    
    # Load valid countries from schema
    valid_countries_schema = load_valid_countries_from_schema()
    current_app.logger.info(f"[push] Loaded {len(valid_countries_schema)} valid countries from schema")
    
    # Cache for country lookups to avoid repeated queries
    country_cache = {}
    
    def create_country_if_missing(country_name):
        """Create a country if it's valid but doesn't exist in Dgraph"""
        try:
            # Try to create the country
            mut = """
            mutation ($in: [AddMemberCountryInput!]!) {
              addMemberCountry(input: $in) {
                memberCountry { countryID title }
              }
            }
            """
            v = {"in": [{"title": country_name}]}
            resp = requests.post(url, json={"query": mut, "variables": v}, headers=headers)
            resp_json = resp.json() if resp else {}
            if resp_json is None:
                resp_json = {}
            
            if "errors" in resp_json and resp_json["errors"]:
                errors_list = resp_json["errors"] if isinstance(resp_json["errors"], list) else [resp_json["errors"]]
                if errors_list[0] is not None and isinstance(errors_list[0], dict):
                    error_msg = errors_list[0].get("message", "Unknown Dgraph error")
                else:
                    error_msg = str(errors_list[0]) if errors_list[0] is not None else "Unknown Dgraph error"
                current_app.logger.error(f"[push] Failed to create country '{country_name}': {error_msg}")
                return None
                
            data = resp_json.get("data", {})
            if data and data.get("addMemberCountry", {}).get("memberCountry"):
                country = data["addMemberCountry"]["memberCountry"][0]
                current_app.logger.info(f"[push] Created new country '{country_name}' with ID: {country['countryID']}")
                return {"countryID": country["countryID"]}
            else:
                current_app.logger.error(f"[push] Unexpected response creating country '{country_name}': {resp_json}")
                return None
                
        except Exception as e:
            current_app.logger.error(f"[push] Exception creating country '{country_name}': {e}")
            # Check if it's a daily limit error
            if "daily limit" in str(e).lower():
                return {"error": f"Daily limit reached: {e}"}
            return None

    def lookup_ref(ref_type_query, var_name, title, id_field):
        # For countries, check cache first
        if ref_type_query == "queryMemberCountry":
            cache_key = f"country_{title}"
            if cache_key in country_cache:
                current_app.logger.debug(f"[push] Found country '{title}' in cache")
                return country_cache[cache_key]
        
        # For countries, use the correct GraphQL query format
        if ref_type_query == "queryMemberCountry":
            q = f'''
            query ($title: String!) {{
              queryMemberCountry(filter: {{title: {{eq: $title}}}}) {{
                {id_field}
              }}
            }}
            '''
        else:
            q = f'''
            query ($title: String!) {{
              {ref_type_query}(filter: {{title: {{eq: $title}}}}) {{
                {id_field}
              }}
            }}
            '''
        current_app.logger.info(f"[push] Looking up {ref_type_query} for title='{title}' ({id_field})")
        try:
            resp = dgraph_request_with_retry(url, {"query": q, "variables": {"title": title}}, headers=headers)
            
            # Validate response structure
            resp_json = resp.json() if resp else {}
            if resp_json is None:
                resp_json = {}
            if "errors" in resp_json and resp_json["errors"]:
                errors_list = resp_json["errors"] if isinstance(resp_json["errors"], list) else [resp_json["errors"]]
                if errors_list[0] is not None and isinstance(errors_list[0], dict):
                    error_msg = errors_list[0].get("message", "Unknown Dgraph error")
                else:
                    error_msg = str(errors_list[0]) if errors_list[0] is not None else "Unknown Dgraph error"
                current_app.logger.error(f"[push] Dgraph query error for {ref_type_query}: {error_msg}")
                # Return a special value to indicate Dgraph error vs "not found"
                return {"error": error_msg}
                
            data = resp_json.get("data", {})
            if not data:
                current_app.logger.warning(f"[push] Empty response data for {ref_type_query}")
                return None
                
            result_list = data.get(ref_type_query, [])
            if not result_list:
                current_app.logger.warning(f"[push] No {ref_type_query} found for '{title}'")
                return None
                
            if id_field not in result_list[0]:
                current_app.logger.error(f"[push] Missing {id_field} in {ref_type_query} response")
                return None
                
            current_app.logger.info(f"[push] Found {id_field} for '{title}': {result_list[0][id_field]}")
            result = {id_field: result_list[0][id_field]}
            
            # Cache country results
            if ref_type_query == "queryMemberCountry":
                cache_key = f"country_{title}"
                country_cache[cache_key] = result
                current_app.logger.debug(f"[push] Cached country '{title}' result")
            
            return result
            
        except Exception as e:
            current_app.logger.error(f"[push] Error for {ref_type_query}: {e}")
            return {"error": str(e)}



    # --- Begin atomic block per company ---
    for m in members:
        biz = m.name or "(Unknown)"
        try:
            # Use Python try/except to make each company atomic (skip if any error occurs)
            # Country lookup: required for every member
            if not m.country1:
                results["errors"].append({
                    "type": "validation_error",
                    "message": f"Missing country for business '{biz}'—skipped.",
                    "business": biz,
                    "field": "country1",
                    "timestamp": datetime.now().isoformat()
                })
                current_app.logger.warning(f"[push] Skipping '{biz}' due to missing country.")
                continue
            # Step 1: Check if country is valid according to schema
            if m.country1 not in valid_countries_schema:
                results["errors"].append({
                    "type": "validation_error",
                    "message": f"Country '{m.country1}' is not a valid country in the schema for business '{biz}'—skipped.",
                    "business": biz,
                    "field": "country1",
                    "value": m.country1,
                    "timestamp": datetime.now().isoformat()
                })
                current_app.logger.warning(f"[push] Skipping '{biz}' due to invalid country '{m.country1}'")
                continue
            
            # Step 2: Check if country exists in Dgraph
            try:
                current_app.logger.debug(f"[push] Looking up country '{m.country1}' for '{biz}'")
                country_ref = lookup_ref("queryMemberCountry", "country", m.country1, "countryID")
                current_app.logger.debug(f"[push] Country lookup result for '{biz}': {country_ref} (type: {type(country_ref)})")
            except Exception as e:
                current_app.logger.error(f"[push] ERROR looking up country '{m.country1}' for '{biz}': {e}", exc_info=True)
                results["errors"].append({
                    "type": "application_error",
                    "message": f"Failed to lookup country '{m.country1}' for business '{biz}'—skipped.",
                    "business": biz,
                    "field": "country1",
                    "value": m.country1,
                    "error_details": str(e),
                    "timestamp": datetime.now().isoformat()
                })
                current_app.logger.warning(f"[push] Skipping '{biz}' due to country lookup error")
                continue
            if country_ref is None:
                # Country not found in Dgraph - try to create it
                current_app.logger.info(f"[push] Country '{m.country1}' not found in Dgraph, attempting to create...")
                country_ref = create_country_if_missing(m.country1)
                if not country_ref:
                    results["errors"].append({
                        "type": "dgraph_error",
                        "message": f"Failed to create country '{m.country1}' for business '{biz}'—skipped.",
                        "business": biz,
                        "field": "country1",
                        "value": m.country1,
                        "timestamp": datetime.now().isoformat()
                    })
                    current_app.logger.warning(f"[push] Skipping '{biz}' due to failed country creation '{m.country1}'")
                    continue
                elif isinstance(country_ref, dict) and "error" in country_ref:
                    # Dgraph error occurred during creation
                    error_msg = country_ref["error"]
                    results["errors"].append({
                        "type": "dgraph_error",
                        "message": f"Dgraph error creating country '{m.country1}': {error_msg} - business '{biz}' skipped.",
                        "business": biz,
                        "field": "country1",
                        "value": m.country1,
                        "error_details": error_msg,
                        "timestamp": datetime.now().isoformat()
                    })
                    current_app.logger.warning(f"[push] Skipping '{biz}' due to Dgraph error creating country '{m.country1}': {error_msg}")
                    continue
                else:
                    current_app.logger.info(f"[push] Successfully created country '{m.country1}' in Dgraph")
            elif isinstance(country_ref, dict) and "error" in country_ref:
                # Dgraph error occurred (e.g., daily limit reached)
                error_msg = country_ref["error"]
                results["errors"].append({
                    "type": "dgraph_error",
                    "message": f"Dgraph error for country '{m.country1}': {error_msg} - business '{biz}' skipped.",
                    "business": biz,
                    "field": "country1",
                    "value": m.country1,
                    "error_details": error_msg,
                    "timestamp": datetime.now().isoformat()
                })
                current_app.logger.warning(f"[push] Skipping '{biz}' due to Dgraph error for country '{m.country1}': {error_msg}")
                continue
            else:
                current_app.logger.info(f"[push] Found existing country '{m.country1}' in Dgraph")

            # Lookup in Dgraph for possible upsert
            q = """
            query ($name: String!) {
              queryMember(filter: {businessName: {eq: $name}}) {
                memberID
                products { title productID }
                ingredients { title ingredientID }
              }
            }
            """
            current_app.logger.info(f"[push] Checking if '{biz}' exists in Dgraph…")
            try:
                current_app.logger.debug(f"[push] Sending member existence query for '{biz}'")
                resp = requests.post(url, json={"query": q, "variables": {"name": biz}}, headers=headers)
                current_app.logger.debug(f"[push] Member existence response for '{biz}': status={resp.status_code if resp else 'None'}")
                
                if resp is None:
                    current_app.logger.error(f"[push] Member existence response is None for '{biz}'")
                    raise Exception("Member existence response is None")
                
                resp_json = resp.json() if resp else {}
                current_app.logger.debug(f"[push] Member existence JSON for '{biz}': {resp_json} (type: {type(resp_json)})")
                
                if resp_json is None:
                    current_app.logger.warning(f"[push] Member existence JSON is None for '{biz}', setting to empty dict")
                    resp_json = {}
                
                # Safe navigation through response structure
                data = resp_json.get("data", {}) if isinstance(resp_json, dict) else {}
                current_app.logger.debug(f"[push] Member existence data for '{biz}': {data} (type: {type(data)})")
                
                node_list = data.get("queryMember", []) if isinstance(data, dict) else []
                current_app.logger.debug(f"[push] Member existence node_list for '{biz}': {node_list} (type: {type(node_list)}, length: {len(node_list) if isinstance(node_list, list) else 'N/A'})")
                
            except Exception as e:
                current_app.logger.error(f"[push] ERROR checking if member '{biz}' exists: {e}", exc_info=True)
                results["errors"].append({
                    "type": "application_error",
                    "message": f"Failed to check if member '{biz}' exists in Dgraph—skipped.",
                    "business": biz,
                    "error_details": str(e),
                    "timestamp": datetime.now().isoformat()
                })
                current_app.logger.warning(f"[push] Skipping '{biz}' due to member existence check error")
                continue

            if node_list:
                current_app.logger.info(f"[push] Member '{biz}' exists, updating products/ingredients…")
                try:
                    node = node_list[0]
                    current_app.logger.debug(f"[push] Existing member node for '{biz}': {node} (type: {type(node)})")
                    
                    mem_id = node.get("memberID") if isinstance(node, dict) else None
                    current_app.logger.debug(f"[push] Member ID for '{biz}': {mem_id}")
                    
                    products = node.get("products", []) if isinstance(node, dict) else []
                    current_app.logger.debug(f"[push] Existing products for '{biz}': {products} (type: {type(products)}, length: {len(products) if isinstance(products, list) else 'N/A'})")
                    
                    ingredients = node.get("ingredients", []) if isinstance(node, dict) else []
                    current_app.logger.debug(f"[push] Existing ingredients for '{biz}': {ingredients} (type: {type(ingredients)}, length: {len(ingredients) if isinstance(ingredients, list) else 'N/A'})")
                    
                    exist_ps = {}
                    if isinstance(products, list):
                        for p in products:
                            if isinstance(p, dict) and "title" in p and "productID" in p:
                                exist_ps[p["title"]] = p["productID"]
                            else:
                                current_app.logger.warning(f"[push] Invalid product entry for '{biz}': {p}")
                    
                    exist_is = {}
                    if isinstance(ingredients, list):
                        for i in ingredients:
                            if isinstance(i, dict) and "title" in i and "ingredientID" in i:
                                exist_is[i["title"]] = i["ingredientID"]
                            else:
                                current_app.logger.warning(f"[push] Invalid ingredient entry for '{biz}': {i}")
                    
                    current_app.logger.debug(f"[push] Processed existing products for '{biz}': {exist_ps}")
                    current_app.logger.debug(f"[push] Processed existing ingredients for '{biz}': {exist_is}")
                    
                except Exception as e:
                    current_app.logger.error(f"[push] ERROR processing existing member data for '{biz}': {e}", exc_info=True)
                    results["errors"].append({
                        "type": "application_error",
                        "message": f"Failed to process existing member data for '{biz}'—skipped.",
                        "business": biz,
                        "error_details": str(e),
                        "timestamp": datetime.now().isoformat()
                    })
                    current_app.logger.warning(f"[push] Skipping '{biz}' due to existing member data processing error")
                    continue

                # Get all products and ingredients for this member
                all_products = [ni for ni in m.new_items if ni.type=="product" and not ni.ignored]
                all_ingredients = [ni for ni in m.new_items if ni.type=="ingredient" and not ni.ignored]
                
                # Separate resolved (matched to canonicals) vs unresolved items
                resolved_products = [ni for ni in all_products if ni.resolved and ni.matched_canonical_id]
                unresolved_products = [ni for ni in all_products if not ni.resolved]
                
                resolved_ingredients = [ni for ni in all_ingredients if ni.resolved and ni.matched_canonical_id]
                unresolved_ingredients = [ni for ni in all_ingredients if not ni.resolved]
                
                current_app.logger.info(f"[push] Member '{biz}' update: {len(resolved_products)} resolved products, {len(unresolved_products)} unresolved products")
                current_app.logger.info(f"[push] Member '{biz}' update: {len(resolved_ingredients)} resolved ingredients, {len(unresolved_ingredients)} unresolved ingredients")

                # Collect all product IDs to link (existing + resolved + new)
                all_product_ids = list(exist_ps.values())  # Start with existing
                new_product_names = []
                
                # Add resolved product IDs (handle multiple selections)
                for ni in resolved_products:
                    # Check if this item has multiple canonical selections
                    if hasattr(ni, 'review') and ni.review and ni.review.alternatives:
                        # Get all selected canonical IDs from alternatives
                        selected_canonicals = []
                        for alt in ni.review.alternatives:
                            if isinstance(alt, dict) and alt.get('selected'):
                                selected_canonicals.append(alt.get('ext_id'))
                        
                        # Add all selected canonical IDs
                        for canonical_id in selected_canonicals:
                            if canonical_id and canonical_id not in all_product_ids:
                                all_product_ids.append(canonical_id)
                                current_app.logger.info(f"[push] Adding multi-selected product '{ni.name}' (ID: {canonical_id}) to existing member '{biz}'")
                    else:
                        # Single selection (backward compatibility)
                        if ni.matched_canonical_id and ni.matched_canonical_id not in all_product_ids:
                            all_product_ids.append(ni.matched_canonical_id)
                            current_app.logger.info(f"[push] Adding resolved product '{ni.name}' (ID: {ni.matched_canonical_id}) to existing member '{biz}'")
                
                # Check unresolved products
                for ni in unresolved_products:
                    if ni.name in exist_ps:
                        # Already linked to this member
                        current_app.logger.info(f"[push] Product '{ni.name}' already linked to member '{biz}'")
                    else:
                        # Check if it exists in Dgraph
                        q = """
                        query ($title: String!) {
                          queryProduct(filter: {title: {eq: $title}}) {
                            productID
                            title
                          }
                        }
                        """
                        resp = requests.post(url, json={"query": q, "variables": {"title": ni.name}}, headers=headers)
                        resp_json = resp.json() if resp else {}
                        if resp_json is None:
                            resp_json = {}
                        existing_products = resp_json.get("data", {}).get("queryProduct", [])
                        
                        if existing_products:
                            # Product exists in Dgraph - link it
                            product_id = existing_products[0]["productID"]
                            if product_id not in all_product_ids:
                                all_product_ids.append(product_id)
                                current_app.logger.info(f"[push] Linking existing product '{ni.name}' (ID: {product_id}) to member '{biz}'")
                        else:
                            # Product doesn't exist - will create new one
                            new_product_names.append(ni.name)
                            current_app.logger.info(f"[push] Will create new product '{ni.name}' for existing member '{biz}'")

                # Same logic for ingredients
                all_ingredient_ids = list(exist_is.values())  # Start with existing
                new_ingredient_names = []
                
                # Add resolved ingredient IDs (handle multiple selections)
                for ni in resolved_ingredients:
                    # Check if this item has multiple canonical selections
                    if hasattr(ni, 'review') and ni.review and ni.review.alternatives:
                        # Get all selected canonical IDs from alternatives
                        selected_canonicals = []
                        for alt in ni.review.alternatives:
                            if isinstance(alt, dict) and alt.get('selected'):
                                selected_canonicals.append(alt.get('ext_id'))
                        
                        # Add all selected canonical IDs
                        for canonical_id in selected_canonicals:
                            if canonical_id and canonical_id not in all_ingredient_ids:
                                all_ingredient_ids.append(canonical_id)
                                current_app.logger.info(f"[push] Adding multi-selected ingredient '{ni.name}' (ID: {canonical_id}) to existing member '{biz}'")
                    else:
                        # Single selection (backward compatibility)
                        if ni.matched_canonical_id and ni.matched_canonical_id not in all_ingredient_ids:
                            all_ingredient_ids.append(ni.matched_canonical_id)
                            current_app.logger.info(f"[push] Adding resolved ingredient '{ni.name}' (ID: {ni.matched_canonical_id}) to existing member '{biz}'")
                
                # Check unresolved ingredients
                for ni in unresolved_ingredients:
                    if ni.name in exist_is:
                        # Already linked to this member
                        current_app.logger.info(f"[push] Ingredient '{ni.name}' already linked to member '{biz}'")
                    else:
                        # Check if it exists in Dgraph
                        q = """
                        query ($title: String!) {
                          queryIngredients(filter: {title: {eq: $title}}) {
                            ingredientID
                            title
                          }
                        }
                        """
                        resp = requests.post(url, json={"query": q, "variables": {"title": ni.name}}, headers=headers)
                        resp_json = resp.json() if resp else {}
                        if resp_json is None:
                            resp_json = {}
                        existing_ingredients = resp_json.get("data", {}).get("queryIngredients", [])
                        
                        if existing_ingredients:
                            # Ingredient exists in Dgraph - link it
                            ingredient_id = existing_ingredients[0]["ingredientID"]
                            if ingredient_id not in all_ingredient_ids:
                                all_ingredient_ids.append(ingredient_id)
                                current_app.logger.info(f"[push] Linking existing ingredient '{ni.name}' (ID: {ingredient_id}) to member '{biz}'")
                        else:
                            # Ingredient doesn't exist - will create new one
                            new_ingredient_names.append(ni.name)
                            current_app.logger.info(f"[push] Will create new ingredient '{ni.name}' for existing member '{biz}'")

                # Create new products if needed
                new_product_ids = []
                if new_product_names:
                    mut = """
                    mutation ($in: [AddProductInput!]!) {
                      addProduct(input: $in) { product { title productID } }
                    }
                    """
                    v = {"in": [{"title": t} for t in new_product_names]}
                    r = requests.post(url, json={"query": mut, "variables": v}, headers=headers)
                    r_json = r.json() if r else {}
                    if r_json is None:
                        r_json = {}
                    if not isinstance(r_json, dict):
                        r_json = {}
                    arr = r_json.get("data", {}).get("addProduct", {}).get("product", [])
                    for pr in arr:
                        new_product_ids.append(pr["productID"])
                        # Add member association note
                        pr["note"] = f"Created with existing member '{biz}'"
                        results["products"].append(pr)
                    current_app.logger.info(f"[push] Created {len(new_product_ids)} new products for existing member '{biz}'")

                # Create new ingredients if needed
                new_ingredient_ids = []
                if new_ingredient_names:
                    mut = """
                    mutation ($in: [AddIngredientsInput!]!) {
                      addIngredients(input: $in) { ingredients { title ingredientID } }
                    }
                    """
                    v = {"in": [{"title": t} for t in new_ingredient_names]}
                    r = requests.post(url, json={"query": mut, "variables": v}, headers=headers)
                    r_json = r.json() if r else {}
                    if r_json is None:
                        r_json = {}
                    if not isinstance(r_json, dict):
                        r_json = {}
                    arr = r_json.get("data", {}).get("addIngredients", {}).get("ingredients", [])
                    for ing in arr:
                        new_ingredient_ids.append(ing["ingredientID"])
                        # Add member association note
                        ing["note"] = f"Created with existing member '{biz}'"
                        results["ingredients"].append(ing)
                    current_app.logger.info(f"[push] Created {len(new_ingredient_ids)} new ingredients for existing member '{biz}'")

                # Update member with all product and ingredient IDs
                all_product_ids.extend(new_product_ids)
                all_ingredient_ids.extend(new_ingredient_ids)
                
                # Add member offerings for existing members
                member_offerings = get_member_offerings_from_cache(m.id)
                offering_refs = []
                if member_offerings:
                    for offering in member_offerings:
                        if isinstance(offering, dict) and offering is not None and 'uid' in offering:
                            offering_refs.append({"offeringID": offering['uid']})
                    if offering_refs:
                        current_app.logger.info(f"[push] Adding {len(offering_refs)} member offerings for existing member '{biz}': {[o.get('title', 'Unknown') for o in member_offerings if isinstance(o, dict) and o is not None]}")
                
                if all_product_ids or all_ingredient_ids or offering_refs:
                    mut = """
                    mutation ($in: UpdateMemberInput!) {
                      updateMember(input: $in) {
                        member { memberID businessName }
                      }
                    }
                    """
                    all_ps = [{"productID": pid} for pid in all_product_ids]
                    all_is = [{"ingredientID": iid} for iid in all_ingredient_ids]
                    
                    update_data = {}
                    if all_ps:
                        update_data["products"] = all_ps
                    if all_is:
                        update_data["ingredients"] = all_is
                    if offering_refs:
                        update_data["memberOfferings"] = offering_refs
                    
                    v = {
                      "in": {
                        "filter": {"memberID": [mem_id]},
                        "set": update_data
                      }
                    }
                    requests.post(url, json={"query": mut, "variables": v}, headers=headers)
                    results["members"].append({"memberID": mem_id, "businessName": biz})
                    current_app.logger.info(f"[push] Updated member '{biz}' with {len(all_product_ids)} products, {len(all_ingredient_ids)} ingredients, and {len(offering_refs)} offerings")

                continue  # done with this existing member

            # 2. Brand-new company → build input
            current_app.logger.info(f"[push] Member '{biz}' is new, creating new record in Dgraph…")
            
            # Get all products and ingredients for this member
            all_products = [ni for ni in m.new_items if ni.type=="product" and not ni.ignored]
            all_ingredients = [ni for ni in m.new_items if ni.type=="ingredient" and not ni.ignored]
            
            # Separate resolved (matched to canonicals) vs unresolved items
            resolved_products = [ni for ni in all_products if ni.resolved and ni.matched_canonical_id]
            unresolved_products = [ni for ni in all_products if not ni.resolved]
            
            resolved_ingredients = [ni for ni in all_ingredients if ni.resolved and ni.matched_canonical_id]
            unresolved_ingredients = [ni for ni in all_ingredients if not ni.resolved]
            
            current_app.logger.info(f"[push] Member '{biz}': {len(resolved_products)} resolved products, {len(unresolved_products)} unresolved products")
            current_app.logger.info(f"[push] Member '{biz}': {len(resolved_ingredients)} resolved ingredients, {len(unresolved_ingredients)} unresolved ingredients")
            
            # For unresolved items, check if they already exist in Dgraph
            existing_product_ids = []
            new_product_names = []
            
            for ni in unresolved_products:
                # Check if this product already exists in Dgraph
                q = """
                query ($title: String!) {
                  queryProduct(filter: {title: {eq: $title}}) {
                    productID
                    title
                  }
                }
                """
                try:
                    resp = requests.post(url, json={"query": q, "variables": {"title": ni.name}}, headers=headers)
                    resp_json = resp.json() if resp else {}
                    if resp_json is None:
                        resp_json = {}
                    existing_products = resp_json.get("data", {}).get("queryProduct", [])
                except Exception as e:
                    current_app.logger.warning(f"[push] Error checking if product '{ni.name}' exists for '{biz}': {e}")
                    # Continue without this product
                    continue
                
                if existing_products:
                    # Product already exists - reuse it
                    existing_product_ids.append(existing_products[0]["productID"])
                    current_app.logger.info(f"[push] Reusing existing product '{ni.name}' (ID: {existing_products[0]['productID']}) for member '{biz}'")
                else:
                    # Product doesn't exist - will create new one
                    new_product_names.append(ni.name)
                    current_app.logger.info(f"[push] Will create new product '{ni.name}' for member '{biz}'")
            
            # Add resolved product IDs (handle multiple selections)
            for ni in resolved_products:
                # Check if this item has multiple canonical selections
                if hasattr(ni, 'review') and ni.review and ni.review.alternatives:
                    # Get all selected canonical IDs from alternatives
                    selected_canonicals = []
                    for alt in ni.review.alternatives:
                        if isinstance(alt, dict) and alt.get('selected'):
                            selected_canonicals.append(alt.get('ext_id'))
                    
                    # Add all selected canonical IDs
                    for canonical_id in selected_canonicals:
                        if canonical_id and canonical_id not in existing_product_ids:
                            existing_product_ids.append(canonical_id)
                            current_app.logger.info(f"[push] Using multi-selected product '{ni.name}' (ID: {canonical_id}) for member '{biz}'")
                else:
                    # Single selection (backward compatibility)
                    if ni.matched_canonical_id and ni.matched_canonical_id not in existing_product_ids:
                        existing_product_ids.append(ni.matched_canonical_id)
                        current_app.logger.info(f"[push] Using resolved product '{ni.name}' (ID: {ni.matched_canonical_id}) for member '{biz}'")
            
            # Same logic for ingredients
            existing_ingredient_ids = []
            new_ingredient_names = []
            
            for ni in unresolved_ingredients:
                # Check if this ingredient already exists in Dgraph
                q = """
                query ($title: String!) {
                  queryIngredients(filter: {title: {eq: $title}}) {
                    ingredientID
                    title
                  }
                }
                """
                try:
                    resp = requests.post(url, json={"query": q, "variables": {"title": ni.name}}, headers=headers)
                    resp_json = resp.json() if resp else {}
                    if resp_json is None:
                        resp_json = {}
                    existing_ingredients = resp_json.get("data", {}).get("queryIngredients", [])
                except Exception as e:
                    current_app.logger.warning(f"[push] Error checking if ingredient '{ni.name}' exists for '{biz}': {e}")
                    # Continue without this ingredient
                    continue
                
                if existing_ingredients:
                    # Ingredient already exists - reuse it
                    existing_ingredient_ids.append(existing_ingredients[0]["ingredientID"])
                    current_app.logger.info(f"[push] Reusing existing ingredient '{ni.name}' (ID: {existing_ingredients[0]['ingredientID']}) for member '{biz}'")
                else:
                    # Ingredient doesn't exist - will create new one
                    new_ingredient_names.append(ni.name)
                    current_app.logger.info(f"[push] Will create new ingredient '{ni.name}' for member '{biz}'")
            
            # Add resolved ingredient IDs (handle multiple selections)
            for ni in resolved_ingredients:
                # Check if this item has multiple canonical selections
                if hasattr(ni, 'review') and ni.review and ni.review.alternatives:
                    # Get all selected canonical IDs from alternatives
                    selected_canonicals = []
                    for alt in ni.review.alternatives:
                        if isinstance(alt, dict) and alt.get('selected'):
                            selected_canonicals.append(alt.get('ext_id'))
                    
                    # Add all selected canonical IDs
                    for canonical_id in selected_canonicals:
                        if canonical_id and canonical_id not in existing_ingredient_ids:
                            existing_ingredient_ids.append(canonical_id)
                            current_app.logger.info(f"[push] Using multi-selected ingredient '{ni.name}' (ID: {canonical_id}) for member '{biz}'")
                else:
                    # Single selection (backward compatibility)
                    if ni.matched_canonical_id and ni.matched_canonical_id not in existing_ingredient_ids:
                        existing_ingredient_ids.append(ni.matched_canonical_id)
                        current_app.logger.info(f"[push] Using resolved ingredient '{ni.name}' (ID: {ni.matched_canonical_id}) for member '{biz}'")
            
            # Create new products if needed
            new_product_ids = []
            if new_product_names:
                mut = """
                mutation ($in: [AddProductInput!]!) {
                  addProduct(input: $in) { product { title productID } }
                }
                """
                v = {"in": [{"title": t} for t in new_product_names]}
                try:
                    r = requests.post(url, json={"query": mut, "variables": v}, headers=headers)
                    r_json = r.json() if r else {}
                    if r_json is None:
                        r_json = {}
                    if not isinstance(r_json, dict):
                        r_json = {}
                    arr = r_json.get("data", {}).get("addProduct", {}).get("product", [])
                except Exception as e:
                    current_app.logger.warning(f"[push] Error creating products for '{biz}': {e}")
                    # Continue without new products
                    arr = []
                for pr in arr:
                    new_product_ids.append(pr["productID"])
                    # Add member association note to avoid duplication
                    pr["note"] = f"Created with member '{biz}'"
                    results["products"].append(pr)
                current_app.logger.info(f"[push] Created {len(new_product_ids)} new products for member '{biz}'")
            
            # Create new ingredients if needed
            new_ingredient_ids = []
            if new_ingredient_names:
                mut = """
                mutation ($in: [AddIngredientsInput!]!) {
                  addIngredients(input: $in) { ingredients { title ingredientID } }
                }
                """
                v = {"in": [{"title": t} for t in new_ingredient_names]}
                try:
                    r = requests.post(url, json={"query": mut, "variables": v}, headers=headers)
                    r_json = r.json() if r else {}
                    if r_json is None:
                        r_json = {}
                    if not isinstance(r_json, dict):
                        r_json = {}
                    arr = r_json.get("data", {}).get("addIngredients", {}).get("ingredients", [])
                except Exception as e:
                    current_app.logger.warning(f"[push] Error creating ingredients for '{biz}': {e}")
                    # Continue without new ingredients
                    arr = []
                for ing in arr:
                    new_ingredient_ids.append(ing["ingredientID"])
                    # Add member association note to avoid duplication
                    ing["note"] = f"Created with member '{biz}'"
                    results["ingredients"].append(ing)
                current_app.logger.info(f"[push] Created {len(new_ingredient_ids)} new ingredients for member '{biz}'")

            state_ref = None
            if hasattr(m, 'state1') and m.state1:
                try:
                    state_ref = lookup_ref("queryMemberStateOrProvince", "state", m.state1, "stateOrProvinceID")
                except Exception as e:
                    current_app.logger.warning(f"[push] Error looking up state '{m.state1}' for '{biz}': {e}")
                    # Continue without state

            # Build member input with validation to ensure no None values
            current_app.logger.debug(f"[push] Building member input for '{biz}'")
            current_app.logger.debug(f"[push] Country ref for '{biz}': {country_ref} (type: {type(country_ref)})")
            current_app.logger.debug(f"[push] Street address for '{biz}': {m.street_address1} (type: {type(m.street_address1)})")
            
            member_input = {
                "businessName":   biz,
                "country1":       country_ref,
                "streetAddress1": m.street_address1 if (m.street_address1 and m.street_address1.strip()) else "Not provided",  # Required field
            }
            current_app.logger.debug(f"[push] Initial member input for '{biz}': {member_input}")
            
            # Only add optional fields if they have valid values
            if m.contact_email and m.contact_email.strip():
                member_input["contactEmail"] = m.contact_email
            if m.city1 and m.city1.strip():
                member_input["city1"] = m.city1
            if m.company_bio and m.company_bio.strip():
                member_input["companyBio"] = m.company_bio
                
            # Add products and ingredients - combine existing and new IDs
            current_app.logger.debug(f"[push] Combining product IDs for '{biz}': existing={existing_product_ids}, new={new_product_ids}")
            current_app.logger.debug(f"[push] Combining ingredient IDs for '{biz}': existing={existing_ingredient_ids}, new={new_ingredient_ids}")
            
            all_product_ids = existing_product_ids + new_product_ids
            all_ingredient_ids = existing_ingredient_ids + new_ingredient_ids
            
            current_app.logger.debug(f"[push] Final product IDs for '{biz}': {all_product_ids} (length: {len(all_product_ids)})")
            current_app.logger.debug(f"[push] Final ingredient IDs for '{biz}': {all_ingredient_ids} (length: {len(all_ingredient_ids)})")
            
            if all_product_ids:
                member_input["products"] = [{"productID": pid} for pid in all_product_ids]
                current_app.logger.debug(f"[push] Added products to member input for '{biz}': {member_input['products']}")
            if all_ingredient_ids:
                member_input["ingredients"] = [{"ingredientID": iid} for iid in all_ingredient_ids]
                current_app.logger.debug(f"[push] Added ingredients to member input for '{biz}': {member_input['ingredients']}")
            if state_ref:
                member_input["stateOrProvince1"] = state_ref
            if hasattr(m, 'zip_code1') and m.zip_code1:
                member_input["zipCode1"] = m.zip_code1
            
            # Add member offerings
            try:
                current_app.logger.debug(f"[push] Getting member offerings for member ID {m.id} (business: '{biz}')")
                member_offerings = get_member_offerings_from_cache(m.id)
                current_app.logger.debug(f"[push] Retrieved member offerings: {member_offerings} (type: {type(member_offerings)})")
                
                if member_offerings:
                    offering_refs = []
                    for i, offering in enumerate(member_offerings):
                        current_app.logger.debug(f"[push] Processing offering {i}: {offering} (type: {type(offering)})")
                        if isinstance(offering, dict) and offering is not None and 'uid' in offering:
                            offering_refs.append({"offeringID": offering['uid']})
                            current_app.logger.debug(f"[push] Added offering ref: {offering['uid']}")
                        else:
                            current_app.logger.warning(f"[push] Skipping invalid offering {i}: {offering} (isinstance dict: {isinstance(offering, dict)}, is None: {offering is None}, has uid: {'uid' in offering if isinstance(offering, dict) else False})")
                    
                    if offering_refs:
                        member_input["memberOfferings"] = offering_refs
                        current_app.logger.info(f"[push] Adding {len(offering_refs)} member offerings for '{biz}': {[o.get('title', 'Unknown') for o in member_offerings if isinstance(o, dict) and o is not None]}")
                    else:
                        current_app.logger.warning(f"[push] No valid offering refs created for '{biz}'")
                else:
                    current_app.logger.debug(f"[push] No member offerings found for '{biz}'")
            except Exception as e:
                current_app.logger.error(f"[push] ERROR getting member offerings for '{biz}': {e}", exc_info=True)
                # Continue without offerings

            mut = """
            mutation ($in: [AddMemberInput!]!) {
              addMember(input: $in) {
                member { memberID businessName }
              }
            }
            """
            current_app.logger.info(f"[push] Final member input for '{biz}': {member_input}")
            current_app.logger.debug(f"[push] Member input type check for '{biz}': {type(member_input)}")
            
            # Validate that all required fields are present and not None
            try:
                if not member_input.get("businessName"):
                    current_app.logger.error(f"[push] Missing businessName in member input for '{biz}'")
                if not member_input.get("country1"):
                    current_app.logger.error(f"[push] Missing country1 in member input for '{biz}'")
                if not member_input.get("streetAddress1"):
                    current_app.logger.error(f"[push] Missing streetAddress1 in member input for '{biz}'")
                
                # Check for None values in the input
                for key, value in member_input.items():
                    if value is None:
                        current_app.logger.error(f"[push] None value found in member input for '{biz}': {key} = {value}")
            except Exception as e:
                current_app.logger.error(f"[push] ERROR validating member input for '{biz}': {e}", exc_info=True)
            try:
                current_app.logger.debug(f"[push] Sending mutation request for '{biz}' to {url}")
                r = requests.post(
                    url,
                    json={"query": mut, "variables": {"in": [member_input]}},
                    headers=headers
                )
                current_app.logger.debug(f"[push] Received response for '{biz}': status={r.status_code if r else 'None'}")
                
                # Detailed response parsing with logging
                if r is None:
                    current_app.logger.error(f"[push] Response object is None for '{biz}'")
                    raise Exception("Response object is None")
                
                resp_json = r.json() if r else {}
                current_app.logger.debug(f"[push] Parsed JSON response for '{biz}': {resp_json} (type: {type(resp_json)})")
                
                if resp_json is None:
                    current_app.logger.warning(f"[push] Response JSON is None for '{biz}', setting to empty dict")
                    resp_json = {}
                if not isinstance(resp_json, dict):
                    current_app.logger.warning(f"[push] Response JSON is not dict for '{biz}': {type(resp_json)}, setting to empty dict")
                    resp_json = {}
                
                # Safe navigation through response structure
                current_app.logger.debug(f"[push] Accessing response data for '{biz}'")
                data = resp_json.get("data", {})
                current_app.logger.debug(f"[push] Response data for '{biz}': {data} (type: {type(data)})")
                
                add_member_data = data.get("addMember", {}) if isinstance(data, dict) else {}
                current_app.logger.debug(f"[push] AddMember data for '{biz}': {add_member_data} (type: {type(add_member_data)})")
                
                arr = add_member_data.get("member", []) if isinstance(add_member_data, dict) else []
                current_app.logger.debug(f"[push] Member array for '{biz}': {arr} (type: {type(arr)}, length: {len(arr) if isinstance(arr, list) else 'N/A'})")
                
            except Exception as e:
                current_app.logger.error(f"[push] ERROR creating member '{biz}': {e}", exc_info=True)
                results["errors"].append({
                    "type": "application_error",
                    "message": f"Failed to create member '{biz}' due to: {e}",
                    "business": biz,
                    "error_details": str(e),
                    "timestamp": datetime.now().isoformat()
                })
                current_app.logger.warning(f"[push] Skipping '{biz}' due to member creation error")
                continue
            if arr:
                results["members"].extend(arr)
                current_app.logger.info(f"[push] Created new member '{biz}' in Dgraph")
                
                # Note: Products and ingredients are already added to results when created above
                # No need to add them again here to avoid duplication
            else:
                current_app.logger.warning(f"[push] No member array returned for '{biz}', checking for errors")
                try:
                    err = resp_json.get("errors", [{"message": "Unknown Dgraph error"}])
                    current_app.logger.debug(f"[push] Error array for '{biz}': {err} (type: {type(err)})")
                    
                    if err and isinstance(err, list) and len(err) > 0:
                        current_app.logger.debug(f"[push] First error for '{biz}': {err[0]} (type: {type(err[0])})")
                        if err[0] is not None and isinstance(err[0], dict):
                            error_msg = err[0].get('message', 'Unknown Dgraph error')
                            current_app.logger.debug(f"[push] Extracted error message for '{biz}': {error_msg}")
                        else:
                            error_msg = str(err[0]) if err[0] is not None else 'Unknown Dgraph error'
                            current_app.logger.debug(f"[push] Converted error to string for '{biz}': {error_msg}")
                    else:
                        error_msg = "Unknown Dgraph error"
                        current_app.logger.debug(f"[push] Using default error message for '{biz}': {error_msg}")
                except Exception as e:
                    current_app.logger.error(f"[push] ERROR processing error response for '{biz}': {e}", exc_info=True)
                    error_msg = f"Error processing response: {e}"
                
                results["errors"].append({
                    "type": "dgraph_error",
                    "message": f"Failed to create '{biz}': {error_msg}",
                    "business": biz,
                    "error_details": error_msg,
                    "timestamp": datetime.now().isoformat()
                })
                current_app.logger.warning(f"[push] Failed to create '{biz}': {error_msg}")
                current_app.logger.warning(f"[push] Full response: {resp_json}")
        except Exception as ex:
            current_app.logger.error(f"[push] ATOMIC ROLLBACK: failed to push '{biz}': {ex}", exc_info=True)
            
            # Check if this is a NoneType error specifically
            if "'NoneType' object has no attribute 'get'" in str(ex):
                current_app.logger.error(f"[push] DETECTED NONETYPE ERROR for '{biz}': {ex}")
                current_app.logger.error(f"[push] This is the specific error we're trying to catch and fix!")
            
            results["errors"].append({
                "type": "application_error",
                "message": f"Failed to push '{biz}' due to: {ex} (skipped; no partial writes for this company)",
                "business": biz,
                "error_details": str(ex),
                "timestamp": datetime.now().isoformat()
            })
            continue

    # Store results in session for downloadable reports
    session['last_push_results'] = results
    session['last_push_errors'] = results["errors"]
    session['last_created_products'] = results["products"]
    session['last_created_ingredients'] = results["ingredients"]
    
    # Log push completion
    logging_manager.log_event("push_complete", {
        "submission_name": submission.name,
        "submission_id": submission.id,
        "operation_id": operation_id,
        "members_created": len(results["members"]),
        "products_created": len(results["products"]),
        "ingredients_created": len(results["ingredients"]),
        "errors_count": len(results["errors"]),
        "success": len(results["errors"]) == 0
    }, operation_id)
    
    return render_template(
        'push_summary.html',
        submission=submission,
        members=results["members"],
        products=results["products"],
        ingredients=results["ingredients"],
        errors=results["errors"],
        operation_id=operation_id
    )

@main_bp.route('/reviews/cancel', methods=['POST'])
def cancel_review():
    current_app.logger.info("[cancel_review] Cancelling review, clearing DB and session.")
    MatchReview.query.delete()
    NewItem.query.delete()
    Member.query.delete()
    MemberSubmission.query.delete()
    db.session.commit()
    session.pop('etl_validation_errors', None)
    session.pop('etl_error_filename', None)
    return redirect(url_for('main.upload_file', status='info', message='Review cancelled. You can upload a new file now.'))

@main_bp.route('/export_results_csv')
def export_results_csv():
    """Export push results as CSV file (legacy route - now redirects to new reports)"""
    submission = MemberSubmission.query.order_by(MemberSubmission.id.desc()).first()
    if not submission:
        current_app.logger.warning("[export_results_csv] No submission found to export.")
        return redirect(url_for('main.upload_file', status='warning', message='No submission found to export.'))
    
    # Redirect to new processed rows report
    return redirect(url_for('main.download_processed_rows_csv', submission_id=submission.id))

@main_bp.route('/download_processed_rows_csv/<int:submission_id>')
def download_processed_rows_csv(submission_id):
    """Download processed_rows.csv report"""
    try:
        csv_content = report_generator.generate_processed_rows_csv(submission_id)
        submission = MemberSubmission.query.get(submission_id)
        filename = f"{submission.name}_processed_rows.csv" if submission else f"submission_{submission_id}_processed_rows.csv"
        
        current_app.logger.info(f"[download_processed_rows_csv] Downloaded processed rows CSV for submission {submission_id}")
        return report_generator.create_csv_response(csv_content, filename)
        
    except Exception as e:
        current_app.logger.error(f"[download_processed_rows_csv] Error: {e}")
        return redirect(url_for('main.upload_file', status='error', message=f'Error generating processed rows CSV: {str(e)}'))

@main_bp.route('/download_errors_csv/<int:submission_id>')
def download_errors_csv(submission_id):
    """Download errors.csv report"""
    try:
        # Get push errors from session or database if available
        push_errors = session.get('last_push_errors', [])
        csv_content = report_generator.generate_errors_csv(submission_id, push_errors)
        submission = MemberSubmission.query.get(submission_id)
        filename = f"{submission.name}_errors.csv" if submission else f"submission_{submission_id}_errors.csv"
        
        current_app.logger.info(f"[download_errors_csv] Downloaded errors CSV for submission {submission_id}")
        return report_generator.create_csv_response(csv_content, filename)
        
    except Exception as e:
        current_app.logger.error(f"[download_errors_csv] Error: {e}")
        return redirect(url_for('main.upload_file', status='error', message=f'Error generating errors CSV: {str(e)}'))

@main_bp.route('/download_created_nodes_csv/<int:submission_id>')
def download_created_nodes_csv(submission_id):
    """Download created_nodes.csv report"""
    try:
        # Get created nodes from session or database if available
        created_products = session.get('last_created_products', [])
        created_ingredients = session.get('last_created_ingredients', [])
        csv_content = report_generator.generate_created_nodes_csv(submission_id, created_products, created_ingredients)
        submission = MemberSubmission.query.get(submission_id)
        filename = f"{submission.name}_created_nodes.csv" if submission else f"submission_{submission_id}_created_nodes.csv"
        
        current_app.logger.info(f"[download_created_nodes_csv] Downloaded created nodes CSV for submission {submission_id}")
        return report_generator.create_csv_response(csv_content, filename)
        
    except Exception as e:
        current_app.logger.error(f"[download_created_nodes_csv] Error: {e}")
        return redirect(url_for('main.upload_file', status='error', message=f'Error generating created nodes CSV: {str(e)}'))

@main_bp.route('/download_all_reports/<int:submission_id>')
def download_all_reports(submission_id):
    """Download all three CSV reports as a zip file"""
    try:
        import zipfile
        import io
        
        # Get push results from session
        push_results = session.get('last_push_results', {})
        reports = report_generator.generate_all_reports(submission_id, push_results)
        
        # Create zip file in memory
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for report_type, csv_content in reports.items():
                zip_file.writestr(f"{report_type}.csv", csv_content)
        
        zip_buffer.seek(0)
        
        # Create response
        submission = MemberSubmission.query.get(submission_id)
        filename = f"{submission.name}_all_reports.zip" if submission else f"submission_{submission_id}_all_reports.zip"
        
        response = make_response(zip_buffer.getvalue())
        response.headers['Content-Type'] = 'application/zip'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        current_app.logger.info(f"[download_all_reports] Downloaded all reports for submission {submission_id}")
        return response
        
    except Exception as e:
        current_app.logger.error(f"[download_all_reports] Error: {e}")
        return redirect(url_for('main.upload_file', status='error', message=f'Error generating all reports: {str(e)}'))

@main_bp.route('/system_status')
def system_status():
    """View system status including daily usage and error statistics"""
    try:
        # Get daily usage info
        daily_usage = error_handler.get_daily_usage_info()
        
        # Get error summary
        error_summary = error_handler.get_error_summary(hours=24)
        
        # Get recent logs
        recent_logs = logging_manager.get_recent_logs(hours=24)
        
        return render_template('system_status.html',
                             daily_usage=daily_usage,
                             error_summary=error_summary,
                             recent_logs=recent_logs[:10])  # Show last 10 logs
        
    except Exception as e:
        current_app.logger.error(f"[system_status] Error: {e}")
        return redirect(url_for('main.upload_file', status='error', message=f'Error loading system status: {str(e)}'))
